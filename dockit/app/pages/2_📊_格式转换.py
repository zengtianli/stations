"""格式转换 — CSV / TXT / XLSX 互转工具."""

import csv
import io

import streamlit as st

from dockit.csv import csv_to_txt, detect_delimiter, txt_to_csv
from dockit.xlsx import csv_to_xlsx, split_sheets, txt_to_xlsx, xls_to_xlsx, xlsx_to_csv

st.set_page_config(page_title="格式转换", page_icon="📊")

st.title("📊 格式转换")
st.markdown("支持 CSV、TXT、XLSX、XLS 之间的格式互转。")

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

ALLOWED_TYPES = ["csv", "txt", "xlsx", "xls"]

# Map source format to available target formats
FORMAT_TARGETS: dict[str, list[str]] = {
    "csv": ["xlsx", "txt"],
    "txt": ["csv", "xlsx"],
    "xlsx": ["csv", "txt", "拆分工作表"],
    "xls": ["xlsx"],
}

DELIMITER_LABELS: dict[str, str] = {
    "\t": "Tab",
    ",": "逗号",
    ";": "分号",
    "|": "竖线",
}


def _detect_format(filename: str) -> str:
    """Return lowercase file extension without dot."""
    return filename.rsplit(".", 1)[-1].lower()


def _preview_csv(csv_string: str, max_rows: int = 10) -> None:
    """Render a CSV string as a table preview and show stats."""
    reader = csv.reader(io.StringIO(csv_string))
    rows = list(reader)
    if not rows:
        st.warning("转换结果为空。")
        return

    total_rows = len(rows)
    total_cols = max(len(r) for r in rows) if rows else 0

    st.info(f"共 **{total_rows}** 行 × **{total_cols}** 列")

    # Use first row as header if it looks like text
    display_rows = rows[: max_rows + 1]
    header = display_rows[0] if display_rows else []
    data = display_rows[1:] if len(display_rows) > 1 else []

    # Pad short rows
    for row in data:
        while len(row) < len(header):
            row.append("")

    st.dataframe(
        data,
        column_config={str(i): h for i, h in enumerate(header)},
        use_container_width=True,
    )
    if total_rows > max_rows + 1:
        st.caption(f"仅显示前 {max_rows} 行数据。")


def _preview_txt(txt_string: str, max_lines: int = 15) -> None:
    """Render plain text preview."""
    lines = txt_string.splitlines()
    st.info(f"共 **{len(lines)}** 行")
    preview = "\n".join(lines[:max_lines])
    st.code(preview, language=None)
    if len(lines) > max_lines:
        st.caption(f"仅显示前 {max_lines} 行。")


def _preview_xlsx_bytes(xlsx_bytes: bytes, label: str = "") -> None:
    """Preview XLSX bytes by converting to CSV internally."""
    csv_dict = xlsx_to_csv(xlsx_bytes)
    for name, csv_str in csv_dict.items():
        title = f"工作表: {name}" if not label else f"{label} — {name}"
        st.markdown(f"**{title}**")
        _preview_csv(csv_str)


def _get_sheet_names(xlsx_bytes: bytes) -> list[str]:
    """Extract sheet names from XLSX bytes."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def _count_csv_stats(csv_string: str) -> tuple[int, int]:
    """Return (row_count, col_count) for a CSV string."""
    reader = csv.reader(io.StringIO(csv_string))
    rows = list(reader)
    if not rows:
        return 0, 0
    return len(rows), max(len(r) for r in rows)


# ---------------------------------------------------------------------------
# File upload
# ---------------------------------------------------------------------------

uploaded = st.file_uploader(
    "上传文件",
    type=ALLOWED_TYPES,
    help="支持 .csv, .txt, .xlsx, .xls 格式",
)

if not uploaded:
    st.stop()

file_bytes = uploaded.read()
src_format = _detect_format(uploaded.name)
file_stem = uploaded.name.rsplit(".", 1)[0]

st.success(f"已上传: **{uploaded.name}** ({len(file_bytes) / 1024:.1f} KB)，格式: **{src_format.upper()}**")

# ---------------------------------------------------------------------------
# Conversion target selection
# ---------------------------------------------------------------------------

targets = FORMAT_TARGETS.get(src_format, [])
if not targets:
    st.error(f"不支持的格式: {src_format}")
    st.stop()

target = st.selectbox("转换目标", targets, help="选择要转换到的格式")

# ---------------------------------------------------------------------------
# Format-specific options & conversion
# ---------------------------------------------------------------------------

if src_format == "xlsx" and target in ("csv", "txt"):
    sheet_names = _get_sheet_names(file_bytes)

    if len(sheet_names) > 1:
        st.markdown(f"检测到 **{len(sheet_names)}** 个工作表")
        convert_all = st.checkbox("转换全部工作表", value=True)
        if not convert_all:
            selected_sheet = st.selectbox("选择工作表", sheet_names)
        else:
            selected_sheet = None
    else:
        selected_sheet = None
        convert_all = True

    if target == "txt":
        delimiter_key = st.selectbox(
            "输出分隔符",
            list(DELIMITER_LABELS.keys()),
            format_func=lambda d: DELIMITER_LABELS[d],
            index=0,
        )

elif src_format == "xlsx" and target == "拆分工作表":
    sheet_names = _get_sheet_names(file_bytes)
    st.markdown(f"检测到 **{len(sheet_names)}** 个工作表: {', '.join(sheet_names)}")

elif src_format == "txt":
    detected = detect_delimiter(file_bytes.decode("utf-8", errors="replace"))
    st.info(f"自动检测分隔符: **{DELIMITER_LABELS.get(detected, repr(detected))}**")
    if target == "xlsx":
        delimiter_key = st.selectbox(
            "确认输入分隔符",
            list(DELIMITER_LABELS.keys()),
            format_func=lambda d: DELIMITER_LABELS[d],
            index=list(DELIMITER_LABELS.keys()).index(detected) if detected in DELIMITER_LABELS else 0,
        )

elif src_format == "csv" and target == "txt":
    delimiter_key = st.selectbox(
        "输出分隔符",
        list(DELIMITER_LABELS.keys()),
        format_func=lambda d: DELIMITER_LABELS[d],
        index=0,
    )

# ---------------------------------------------------------------------------
# Execute conversion
# ---------------------------------------------------------------------------

if st.button("开始转换", type="primary", use_container_width=True):
    try:
        # --- XLSX source ---
        if src_format == "xlsx" and target == "csv":
            csv_dict = xlsx_to_csv(file_bytes, sheet_name=selected_sheet if not convert_all else None)
            st.success(f"转换完成！共 {len(csv_dict)} 个工作表。")

            for name, csv_str in csv_dict.items():
                rows, cols = _count_csv_stats(csv_str)
                st.markdown(f"### 工作表: {name}  ({rows} 行 × {cols} 列)")
                _preview_csv(csv_str)

                download_name = f"{file_stem}_{name}.csv" if len(csv_dict) > 1 else f"{file_stem}.csv"
                st.download_button(
                    f"下载 {name}.csv",
                    data=csv_str.encode("utf-8-sig"),
                    file_name=download_name,
                    mime="text/csv",
                    key=f"dl_csv_{name}",
                )

        elif src_format == "xlsx" and target == "txt":
            csv_dict = xlsx_to_csv(file_bytes, sheet_name=selected_sheet if not convert_all else None)
            st.success(f"转换完成！共 {len(csv_dict)} 个工作表。")

            for name, csv_str in csv_dict.items():
                txt_str = csv_to_txt(csv_str, delimiter=delimiter_key)
                _preview_txt(txt_str)

                download_name = f"{file_stem}_{name}.txt" if len(csv_dict) > 1 else f"{file_stem}.txt"
                st.download_button(
                    f"下载 {name}.txt",
                    data=txt_str.encode("utf-8"),
                    file_name=download_name,
                    mime="text/plain",
                    key=f"dl_txt_{name}",
                )

        elif src_format == "xlsx" and target == "拆分工作表":
            sheets = split_sheets(file_bytes)
            st.success(f"拆分完成！共 {len(sheets)} 个工作表文件。")

            for name, sheet_bytes in sheets.items():
                st.markdown(f"### 工作表: {name}")
                _preview_xlsx_bytes(sheet_bytes, label=name)
                st.download_button(
                    f"下载 {name}.xlsx",
                    data=sheet_bytes,
                    file_name=f"{file_stem}_{name}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_sheet_{name}",
                )

        # --- XLS source ---
        elif src_format == "xls" and target == "xlsx":
            result_bytes = xls_to_xlsx(file_bytes)
            st.success("转换完成！XLS → XLSX")
            _preview_xlsx_bytes(result_bytes)
            st.download_button(
                "下载 XLSX",
                data=result_bytes,
                file_name=f"{file_stem}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # --- CSV source ---
        elif src_format == "csv" and target == "xlsx":
            csv_content = file_bytes.decode("utf-8", errors="replace")
            result_bytes = csv_to_xlsx(csv_content)
            rows, cols = _count_csv_stats(csv_content)
            st.success(f"转换完成！{rows} 行 × {cols} 列")
            _preview_csv(csv_content)
            st.download_button(
                "下载 XLSX",
                data=result_bytes,
                file_name=f"{file_stem}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        elif src_format == "csv" and target == "txt":
            csv_content = file_bytes.decode("utf-8", errors="replace")
            txt_str = csv_to_txt(csv_content, delimiter=delimiter_key)
            st.success("转换完成！CSV → TXT")
            _preview_txt(txt_str)
            st.download_button(
                "下载 TXT",
                data=txt_str.encode("utf-8"),
                file_name=f"{file_stem}.txt",
                mime="text/plain",
            )

        # --- TXT source ---
        elif src_format == "txt" and target == "csv":
            txt_content = file_bytes.decode("utf-8", errors="replace")
            csv_str = txt_to_csv(txt_content)
            rows, cols = _count_csv_stats(csv_str)
            st.success(f"转换完成！{rows} 行 × {cols} 列")
            _preview_csv(csv_str)
            st.download_button(
                "下载 CSV",
                data=csv_str.encode("utf-8-sig"),
                file_name=f"{file_stem}.csv",
                mime="text/csv",
            )

        elif src_format == "txt" and target == "xlsx":
            txt_content = file_bytes.decode("utf-8", errors="replace")
            result_bytes = txt_to_xlsx(txt_content, delimiter=delimiter_key)
            st.success("转换完成！TXT → XLSX")
            # Preview via csv conversion
            csv_str = txt_to_csv(txt_content)
            _preview_csv(csv_str)
            st.download_button(
                "下载 XLSX",
                data=result_bytes,
                file_name=f"{file_stem}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        else:
            st.error(f"暂不支持 {src_format.upper()} → {target.upper()} 的转换。")

    except Exception as e:
        st.error(f"转换失败: {e}")
