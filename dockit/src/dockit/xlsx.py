"""Excel processing — conversion, splitting, merging, header manipulation.

Pure logic module: bytes in, bytes out. Uses openpyxl (no pandas required for core ops).

Usage:
    from dockit.xlsx import xlsx_to_csv, csv_to_xlsx, split_sheets

    csv_dict = xlsx_to_csv(xlsx_bytes)           # {sheet_name: csv_string}
    xlsx_bytes = csv_to_xlsx("col1,col2\\n1,2")  # CSV string -> XLSX bytes
    sheets = split_sheets(xlsx_bytes)             # {sheet_name: xlsx_bytes}
"""

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook, load_workbook


def xlsx_to_csv(xlsx_bytes: bytes, sheet_name: str | None = None) -> dict[str, str]:
    """Convert XLSX to CSV strings.

    Args:
        xlsx_bytes: Raw bytes of an .xlsx file.
        sheet_name: Specific sheet to convert. If None, converts all sheets.

    Returns:
        Dict mapping sheet names to CSV strings.
    """
    wb = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    result = {}

    sheets = [sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames

    for name in sheets:
        ws = wb[name]
        buf = StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(c) if c is not None else "" for c in row])
        result[name] = buf.getvalue()

    wb.close()
    return result


def csv_to_xlsx(csv_content: str) -> bytes:
    """Convert a CSV string to XLSX bytes.

    Args:
        csv_content: CSV-formatted string.

    Returns:
        XLSX file as bytes.
    """
    wb = Workbook()
    ws = wb.active

    reader = csv.reader(StringIO(csv_content))
    for row_idx, row in enumerate(reader, 1):
        for col_idx, value in enumerate(row, 1):
            # Try to convert numeric values
            try:
                value = int(value)
            except (ValueError, TypeError):
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def txt_to_xlsx(content: str, delimiter: str = "\t") -> bytes:
    """Convert delimited text to XLSX bytes.

    Args:
        content: Tab-delimited (or other delimiter) text.
        delimiter: Field delimiter.

    Returns:
        XLSX file as bytes.
    """
    wb = Workbook()
    ws = wb.active

    for row_idx, line in enumerate(content.splitlines(), 1):
        if not line.strip():
            continue
        for col_idx, value in enumerate(line.split(delimiter), 1):
            try:
                value = int(value)
            except (ValueError, TypeError):
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass
            ws.cell(row=row_idx, column=col_idx, value=value)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def split_sheets(xlsx_bytes: bytes) -> dict[str, bytes]:
    """Split a workbook into per-sheet XLSX files.

    Args:
        xlsx_bytes: Raw bytes of an .xlsx file.

    Returns:
        Dict mapping sheet names to individual XLSX bytes.
    """
    wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb_new = Workbook()
        ws_new = wb_new.active
        ws_new.title = sheet_name

        for row in ws.iter_rows():
            for cell in row:
                ws_new.cell(row=cell.row, column=cell.column, value=cell.value)

        buf = BytesIO()
        wb_new.save(buf)
        result[sheet_name] = buf.getvalue()

    wb.close()
    return result


def lowercase_headers(xlsx_bytes: bytes) -> bytes:
    """Lowercase all column headers (first row) in a workbook.

    Args:
        xlsx_bytes: Raw bytes of an .xlsx file.

    Returns:
        Modified XLSX file as bytes.
    """
    wb = load_workbook(BytesIO(xlsx_bytes))

    for ws in wb.worksheets:
        for cell in ws[1]:
            if cell.value and isinstance(cell.value, str):
                cell.value = cell.value.lower()

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xls_to_xlsx(xls_bytes: bytes) -> bytes:
    """Convert legacy .xls to .xlsx format.

    Requires xlrd to be installed (`pip install dockit[excel]`).

    Args:
        xls_bytes: Raw bytes of an .xls file.

    Returns:
        XLSX file as bytes.
    """
    import xlrd

    wb_old = xlrd.open_workbook(file_contents=xls_bytes)
    wb_new = Workbook()

    for sheet_name in wb_old.sheet_names():
        ws_old = wb_old.sheet_by_name(sheet_name)
        ws_new = wb_new.create_sheet(title=sheet_name)
        for row in range(ws_old.nrows):
            for col in range(ws_old.ncols):
                ws_new.cell(row=row + 1, column=col + 1, value=ws_old.cell_value(row, col))

    # Remove the default empty sheet
    if "Sheet" in wb_new.sheetnames:
        del wb_new["Sheet"]

    buf = BytesIO()
    wb_new.save(buf)
    return buf.getvalue()
