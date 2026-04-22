#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
解析纳污能力计算的 xlsx 输入文件（竖向布局）

输入表布局（河道功能区-输入）:
  行1: 功能区数量(B1), 功能区编号(E1起)
  行2: 方案个数(B2), 功能区名(E2起)
  行3: Cs
  行4: K(1/s)
  行5: b
  行6: a
  行7: β
  行8: 干流总长L(m)
  行9: 干流C0
  行10: 支流数量
  行11: 干流名
  行12起: 支流信息（每条支流5行：分隔行、名称、长度、汇入位置、浓度）
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Tuple, Optional
from .calc_core import Zone, Branch, ReservoirZone


def parse_input_sheet(ws_data: list) -> Tuple[List[Zone], int]:
    """
    解析河道功能区-输入 sheet（竖向布局）

    Args:
        ws_data: openpyxl worksheet 的行数据列表，每行为 cell value 列表

    Returns:
        (zones, scheme_count): 功能区列表 和 方案个数
    """
    # 行索引（0-based）
    ROW_ZONE_COUNT = 0
    ROW_SCHEME_COUNT = 1
    ROW_CS = 2
    ROW_K = 3
    ROW_B = 4
    ROW_A = 5
    ROW_BETA = 6
    ROW_MAIN_LENGTH = 7
    ROW_MAIN_C0 = 8
    ROW_BRANCH_COUNT = 9
    ROW_MAIN_NAME = 10
    ROW_BRANCH_START = 11
    ROWS_PER_BRANCH = 5  # 分隔行、名称、长度、汇入位置、浓度

    # 数据列从 col=4 (E列, 0-based index=4) 开始
    DATA_COL_START = 4

    zone_count = int(ws_data[ROW_ZONE_COUNT][1])  # B1
    scheme_count = int(ws_data[ROW_SCHEME_COUNT][1])  # B2

    zones = []
    for i in range(zone_count):
        col = DATA_COL_START + i

        zone_id = _safe_str(ws_data[ROW_ZONE_COUNT][col])  # 功能区编号
        name = _safe_str(ws_data[ROW_SCHEME_COUNT][col])  # 功能区名
        Cs = _safe_float(ws_data[ROW_CS][col])
        K = _safe_float(ws_data[ROW_K][col])
        b = _safe_float(ws_data[ROW_B][col])
        a = _safe_float(ws_data[ROW_A][col])
        beta = _safe_float(ws_data[ROW_BETA][col])
        main_length = _safe_float(ws_data[ROW_MAIN_LENGTH][col])
        main_C0 = _safe_float(ws_data[ROW_MAIN_C0][col])
        branch_count = int(_safe_float(ws_data[ROW_BRANCH_COUNT][col]))
        main_name = _safe_str(ws_data[ROW_MAIN_NAME][col])

        # 解析支流
        branches = []
        for j in range(branch_count):
            base_row = ROW_BRANCH_START + j * ROWS_PER_BRANCH
            # 分隔行(+0), 名称(+1), 长度(+2), 汇入位置(+3), 浓度(+4)
            if base_row + 4 < len(ws_data):
                br_name = _safe_str(ws_data[base_row + 1][col])
                br_length = _safe_float(ws_data[base_row + 2][col])
                br_join_pos = _safe_float(ws_data[base_row + 3][col])
                br_C0 = _safe_float(ws_data[base_row + 4][col])
                branches.append(Branch(
                    name=br_name,
                    length=br_length,
                    join_position=br_join_pos,
                    C0=br_C0,
                ))

        zone = Zone(
            zone_id=name,  # 用功能区名作为 ID（与流量表列名对应）
            name=name,
            water_class="",
            length=main_length,
            K=K,
            b=b,
            a=a,
            beta=beta,
            Cs=Cs,
            C0=main_C0,
            main_name=main_name,
            branches=branches,
        )
        zones.append(zone)

    return zones, scheme_count


def parse_flow_sheets(xlsx: pd.ExcelFile, scheme_count: int) -> Dict[int, pd.DataFrame]:
    """
    解析多方案的逐日流量 sheet

    Args:
        xlsx: pd.ExcelFile 对象
        scheme_count: 方案个数

    Returns:
        {方案号(1-based): 逐日流量 DataFrame}
    """
    flows = {}
    for s in range(1, scheme_count + 1):
        sheet_name = _find_sheet(xlsx.sheet_names, f"逐日流量", f"方案{s}")
        if sheet_name:
            df = pd.read_excel(xlsx, sheet_name=sheet_name)
            df['日期'] = pd.to_datetime(df['日期'])
            flows[s] = df
    return flows


def parse_reservoir_input(xlsx: pd.ExcelFile) -> Tuple[List[ReservoirZone], Optional[pd.DataFrame]]:
    """
    解析水库输入（如有）

    Returns:
        (reservoir_zones, daily_volume_df) 或 ([], None)
    """
    zones_sheet = _find_sheet(xlsx.sheet_names, "水库功能区基础信息")
    volume_sheet = _find_sheet(xlsx.sheet_names, "水库逐日库容")

    if not zones_sheet or not volume_sheet:
        return [], None

    zones_df = pd.read_excel(xlsx, sheet_name=zones_sheet)
    zones = []
    for _, row in zones_df.iterrows():
        zone = ReservoirZone(
            zone_id=str(row['功能区']),
            name=str(row['名称']),
            K=float(row['K(1/s)']),
            b=float(row['b']),
            Cs=float(row['Cs']) if pd.notna(row.get('Cs')) else 0.0,
            C0=float(row['C0']) if pd.notna(row.get('C0')) else 0.0,
        )
        zones.append(zone)

    volume_df = pd.read_excel(xlsx, sheet_name=volume_sheet)
    volume_df['日期'] = pd.to_datetime(volume_df['日期'])

    return zones, volume_df


def get_flow_column_map(zones: List[Zone], flow_columns: list) -> Dict[str, Dict[str, str]]:
    """
    建立功能区 → 流量列名的映射

    Returns:
        {zone_id: {"main": "钱塘156-0", "branches": ["钱塘156-1", ...]}}
    """
    col_map = {}
    for zone in zones:
        entry = {"main": zone.main_name, "branches": []}
        for br in zone.branches:
            if br.name in flow_columns:
                entry["branches"].append(br.name)
        col_map[zone.zone_id] = entry
    return col_map


def read_input_sheet_raw(xlsx_file) -> list:
    """
    用 openpyxl 读取输入 sheet 的原始数据（保留竖向布局）

    Args:
        xlsx_file: 文件路径或 file-like 对象

    Returns:
        行列表，每行为 cell value 列表
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)

    # 查找输入 sheet
    input_sheet_name = None
    for name in wb.sheetnames:
        if "输入" in name and "功能区" in name:
            input_sheet_name = name
            break

    if not input_sheet_name:
        raise ValueError("未找到包含'功能区'和'输入'的 sheet")

    ws = wb[input_sheet_name]
    data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        data.append(list(row))
    wb.close()
    return data


def _find_sheet(sheet_names: list, *keywords: str) -> Optional[str]:
    """根据关键字查找 sheet 名"""
    for name in sheet_names:
        if all(kw in name for kw in keywords):
            return name
    return None


def _safe_float(val) -> float:
    if val is None or val == "" or val == "None":
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _safe_str(val) -> str:
    if val is None:
        return ""
    return str(val).strip()
