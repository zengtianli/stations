#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
保护片数据导入脚本 - 从 GeoJSON 导入到 Excel 数据库

【使用方法】
    python3 fill_bhp_to_database.py

【功能】
    1. 从 input/保护片/env.geojson 读取熟溪保护片数据
    2. 从 input/city_county_town.csv 读取乡镇编码
    3. 参考 database_bx.xlsx（白溪模板）的字段结构
    4. 自动转换拼音简写（钱塘江流域→QTJLY，熟溪→SX等）
    5. 统一编码为大写格式（sx0001→SX0001）
    6. 填充到 output/datebase_sx.xlsx 的"保护片" sheet
    7. 自动备份原文件
    8. 生成详细报告

【文件要求】
    - input/保护片/env.geojson       源数据（熟溪保护片GeoJSON）
    - input/city_county_town.csv      乡镇编码对照表
    - templates/database_bx.xlsx      白溪模板（参考字段结构）
    - output/datebase_sx.xlsx         目标Excel（熟溪，必须有"保护片"sheet）

【依赖安装】
    pip install pandas openpyxl

【自定义映射】
    修改脚本第46-57行的映射表：
    RIVER_NAME_TO_CODE    河流名称→编码
    BASIN_NAME_TO_CODE    流域名称→编码
"""

import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    get_river_code,
    get_basin_code,
    normalize_code,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def load_mappings(town_csv_path, region_csv_path):
    """
    从 CSV 加载映射表
    返回: (town_to_county, county_to_code, town_to_code)
    - town_to_county: {乡镇名称: 县名称}
    - county_to_code: {县名称: 县编码}
    - town_to_code: {乡镇名称: 乡镇编码}
    """
    town_to_county = {}
    county_to_code = {}
    town_to_code = {}
    
    # 1. 加载乡镇 → 县映射
    try:
        df_town = pd.read_csv(town_csv_path, encoding='utf-8')
        if 'town' in df_town.columns and 'county' in df_town.columns:
            for _, row in df_town.iterrows():
                town = row.get('town', '')
                county = row.get('county', '')
                if town and county:
                    town_to_county[town] = county
        print(f"✓ 成功加载 {len(town_to_county)} 个乡镇→县映射")
    except Exception as e:
        print(f"⚠ 警告: 无法加载乡镇映射 - {e}")
    
    # 2. 加载县名 → 县编码映射 和 乡镇名 → 乡镇编码映射
    try:
        df_region = pd.read_csv(region_csv_path, encoding='utf-8', dtype={'region_code': str})
        if 'region_name' in df_region.columns and 'region_code' in df_region.columns:
            for _, row in df_region.iterrows():
                region_name = row.get('region_name', '')
                region_code = row.get('region_code', '')
                if region_name and region_code:
                    # 清理编码格式（去除科学计数法、小数点等）
                    try:
                        code_int = int(float(region_code))
                        code_str = str(code_int)
                        
                        if len(code_str) == 15:
                            # 县级（后9位是0）
                            if code_str.endswith('000000000'):
                                county_to_code[region_name] = code_str
                            # 乡镇级（倒数第4-6位不全是0，即街道/乡/镇）
                            elif not code_str.endswith('000000000'):
                                town_to_code[region_name] = code_str
                    except:
                        pass
        print(f"✓ 成功加载 {len(county_to_code)} 个县→编码映射")
        print(f"✓ 成功加载 {len(town_to_code)} 个乡镇→编码映射")
    except Exception as e:
        print(f"⚠ 警告: 无法加载编码映射 - {e}")
    
    return town_to_county, county_to_code, town_to_code


def fill_baohuplan_data(geojson_data, excel_path, town_to_county, county_to_code, town_to_code, sheet_name='保护片'):
    """
    将 GeoJSON 数据填充到 Excel 的保护片 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        town_to_county: 乡镇→县映射
        county_to_code: 县→编码映射
        town_to_code: 乡镇→编码映射
        sheet_name: sheet 名称
    """
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'code': 1,                      # 保护片编码
        'name': 2,                      # 保护片名称
        'adcdName': 3,                  # 所属区域（县）
        'adcdCode': 4,                  # 区划编码(县）
        'townName': 5,                  # 所在乡镇
        'townCode': 6,                  # 乡镇编码
        'area': 7,                      # 面积（km2）
        'proPeople': 8,                 # 保护人口（万人）
        'gdp': 9,                       # 年生产总值（万元）
        'pop_pm': 10,                   # 每平方千米人口（万人/km²）
        'gdp_pm': 11,                   # 每平方千米GDP（万元/km²）
        'houseArea': 12,                # 房屋面积（万m2）
        'roadLen': 13,                  # 道路（km）
        'protectLand': 14,              # 保护耕地（公顷）
        'type': 15,                     # 现状标准
        'planType': 16,                 # 规划标准
        'isComplete': 17,               # 达标情况(0不达标1达标)
        'river': 18,                    # 防御河流
        'proObj': 19,                   # 保护对象
        'riverAlias': 20,               # 流域编码（八大流域首字母简写）
        'subCode': 21,                  # 支流编码
        'level': 22,                    # 级别
        'lng': 23,                      # 经度
        'lat': 24,                      # 纬度
        'minGroundElevation': 25,       # 最低地面高程
        'pumpingCapacity': 26,          # 电排规模(m³/s)
        'maxDikeCrestElevation': 27,    # 最大堤顶高程
        'minDikeCrestElevation': 28,    # 最小堤顶高程
        'avgGroundElevation': 29,       # 平均地面高程
        'designWaterLevel': 30,         # 设计水位
        'drainageStandard': 31,         # 排涝标准
        'drainageModulus': 32,          # 排涝模数
        'drainageFlow': 33,             # 排水流量
        'topLeftLongitude': 34,         # 左上角经度
        'topLeftLatitude': 35,          # 左上角纬度
        'bottomRightLongitude': 36,     # 右下角经度
        'bottomRightLatitude': 37,      # 右下角纬度
        'lake_code': 38,                # 关联的概湖编码
        'dot_code': 39,                 # 关联的断面编码
        'geom': 40                      # 面
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据
    start_row = 3
    features = geojson_data.get('features', [])
    
    print(f"✓ 开始填充数据（自动处理拼音简写）...")
    
    # 统计转换信息
    conversions = {
        'code': 0,
        'riverAlias': 0,
        'subCode': 0,
    }
    
    for idx, feature in enumerate(features):
        props = feature.get('properties', {})
        geom = feature.get('geometry', {})
        row_num = start_row + idx
        
        # 保护片编码 - 规范化为大写
        code = normalize_code(props.get('code', ''))
        if code != props.get('code', ''):
            conversions['code'] += 1
        ws.cell(row=row_num, column=column_mapping['code']).value = code
        
        # 基本信息
        ws.cell(row=row_num, column=column_mapping['name']).value = props.get('name', '')
        
        # 所在乡镇 - 从 GeoJSON 的 adcdName 字段获取（可能有多个乡镇）
        town_name = props.get('adcdName', '')  # 注意：GeoJSON 的 adcdName 对应 Excel 的 townName
        ws.cell(row=row_num, column=column_mapping['townName']).value = town_name
        
        # 所属区域（县）- 从第一个乡镇查县名
        first_town = town_name.split('、')[0] if town_name else ''
        adcd_name = town_to_county.get(first_town, '')
        ws.cell(row=row_num, column=column_mapping['adcdName']).value = adcd_name
        
        # 区划编码（县）- 从县名查县编码
        adcd_code = county_to_code.get(adcd_name, '')
        ws.cell(row=row_num, column=column_mapping['adcdCode']).value = adcd_code
        
        # 乡镇编码 - 从多个乡镇分别查询编码，用顿号连接
        town_codes = []
        if town_name:
            for town in town_name.split('、'):
                town = town.strip()
                if town and town in town_to_code:
                    town_codes.append(town_to_code[town])
        town_code = '、'.join(town_codes) if town_codes else ''
        ws.cell(row=row_num, column=column_mapping['townCode']).value = town_code
        
        # 数值字段
        area = props.get('area')
        if area is not None:
            ws.cell(row=row_num, column=column_mapping['area']).value = round(float(area), 6)
        
        proPeople = props.get('proPeople')
        if proPeople is not None:
            ws.cell(row=row_num, column=column_mapping['proPeople']).value = round(float(proPeople), 6)
        
        gdp = props.get('gdp')
        if gdp is not None:
            ws.cell(row=row_num, column=column_mapping['gdp']).value = round(float(gdp), 4)
        
        # 计算密度字段
        if area and area > 0:
            if proPeople is not None:
                pop_pm = float(proPeople) / float(area)
                ws.cell(row=row_num, column=column_mapping['pop_pm']).value = round(pop_pm, 6)
            if gdp is not None:
                gdp_pm = float(gdp) / float(area)
                ws.cell(row=row_num, column=column_mapping['gdp_pm']).value = round(gdp_pm, 4)
        
        houseArea = props.get('houseArea')
        if houseArea is not None:
            ws.cell(row=row_num, column=column_mapping['houseArea']).value = round(float(houseArea), 6)
        
        roadLen = props.get('roadLen')
        if roadLen is not None:
            ws.cell(row=row_num, column=column_mapping['roadLen']).value = round(float(roadLen), 6)
        
        protectLand = props.get('protectLand')
        if protectLand is not None:
            ws.cell(row=row_num, column=column_mapping['protectLand']).value = round(float(protectLand), 6)
        
        # 标准和状态
        ws.cell(row=row_num, column=column_mapping['type']).value = props.get('type', '')
        ws.cell(row=row_num, column=column_mapping['planType']).value = props.get('planType', '')
        ws.cell(row=row_num, column=column_mapping['isComplete']).value = props.get('isComplete', '')
        
        # 河流信息
        ws.cell(row=row_num, column=column_mapping['river']).value = props.get('river', '')
        ws.cell(row=row_num, column=column_mapping['proObj']).value = props.get('proObj', '')
        
        # 流域编码 - 转换为拼音简写（短版本）
        river_alias = props.get('riverAlias', '')
        river_alias_code = get_basin_code(river_alias)
        if river_alias != river_alias_code:
            conversions['riverAlias'] += 1
        ws.cell(row=row_num, column=column_mapping['riverAlias']).value = river_alias_code
        
        # 支流编码 - 从河流名称转为拼音简写
        sub_code_raw = props.get('subCode', '')
        sub_code = get_river_code(sub_code_raw)  # 转为拼音简写（如：熟溪→SX）
        sub_code = normalize_code(sub_code)  # 转为大写
        if sub_code != sub_code_raw:
            conversions['subCode'] += 1
        ws.cell(row=row_num, column=column_mapping['subCode']).value = sub_code
        
        ws.cell(row=row_num, column=column_mapping['level']).value = props.get('level', '')
        
        # 坐标信息
        lng = props.get('lng')
        if lng is not None:
            ws.cell(row=row_num, column=column_mapping['lng']).value = round(float(lng), 6)
        
        lat = props.get('lat')
        if lat is not None:
            ws.cell(row=row_num, column=column_mapping['lat']).value = round(float(lat), 6)
        
        # 高程信息
        ws.cell(row=row_num, column=column_mapping['minGroundElevation']).value = props.get('minGroundElevation', '')
        ws.cell(row=row_num, column=column_mapping['pumpingCapacity']).value = props.get('pumpingCapacity', '')
        ws.cell(row=row_num, column=column_mapping['maxDikeCrestElevation']).value = props.get('maxDikeCrestElevation', '')
        ws.cell(row=row_num, column=column_mapping['minDikeCrestElevation']).value = props.get('minDikeCrestElevation', '')
        ws.cell(row=row_num, column=column_mapping['avgGroundElevation']).value = props.get('avgGroundElevation', '')
        ws.cell(row=row_num, column=column_mapping['designWaterLevel']).value = props.get('designWaterLevel', '')
        
        # 排涝信息
        ws.cell(row=row_num, column=column_mapping['drainageStandard']).value = props.get('drainageStandard', '')
        ws.cell(row=row_num, column=column_mapping['drainageModulus']).value = props.get('drainageModulus', '')
        ws.cell(row=row_num, column=column_mapping['drainageFlow']).value = props.get('drainageFlow', '')
        
        # 边界坐标
        topLeftLongitude = props.get('topLeftLongitude')
        if topLeftLongitude is not None:
            ws.cell(row=row_num, column=column_mapping['topLeftLongitude']).value = round(float(topLeftLongitude), 6)
        
        topLeftLatitude = props.get('topLeftLatitude')
        if topLeftLatitude is not None:
            ws.cell(row=row_num, column=column_mapping['topLeftLatitude']).value = round(float(topLeftLatitude), 6)
        
        bottomRightLongitude = props.get('bottomRightLongitude')
        if bottomRightLongitude is not None:
            ws.cell(row=row_num, column=column_mapping['bottomRightLongitude']).value = round(float(bottomRightLongitude), 6)
        
        bottomRightLatitude = props.get('bottomRightLatitude')
        if bottomRightLatitude is not None:
            ws.cell(row=row_num, column=column_mapping['bottomRightLatitude']).value = round(float(bottomRightLatitude), 6)
        
        # 关联编码和几何信息 - 强制留空
        ws.cell(row=row_num, column=column_mapping['lake_code']).value = ''
        ws.cell(row=row_num, column=column_mapping['dot_code']).value = ''
        ws.cell(row=row_num, column=column_mapping['geom']).value = ''
        
        # 显示进度
        if (idx + 1) % 5 == 0 or idx == len(features) - 1:
            print(f"  已填充 {idx + 1}/{len(features)} 条记录...")
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示转换统计
    print(f"\n✓ 编码转换统计:")
    print(f"  - code 规范化: {conversions['code']} 条（小写 → 大写）")
    print(f"  - riverAlias 转换: {conversions['riverAlias']} 条（流域名称 → 短简写，如 QTJLY → QTJ）")
    print(f"  - subCode 转换: {conversions['subCode']} 条（河流名称 → 拼音简写，如 熟溪 → SX）")
    
    return len(features)


def generate_report(excel_path, geojson_path, filled_count, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='保护片')
    data_df = df.iloc[1:].reset_index(drop=True)  # 跳过表头说明行
    
    report = f"""
{'=' * 80}
保护片数据填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

一、填充统计
{'-' * 80}
总记录数: {filled_count} 条

二、河流分布
{'-' * 80}
{data_df['river'].value_counts().to_string()}

三、所属区域分布
{'-' * 80}
{data_df['adcdName'].value_counts().to_string()}

四、达标情况统计
{'-' * 80}
达标: {len(data_df[data_df['isComplete'] == 1])} 条
不达标: {len(data_df[data_df['isComplete'] == 0])} 条

五、保护对象类型
{'-' * 80}
{data_df['proObj'].value_counts().to_string()}

六、编码规范化说明
{'-' * 80}
- code: 已规范化为大写（如：sx0001 → SX0001）
- adcdName: 已从乡镇查询县名（如：白姆乡 → 武义县）
- adcdCode: 已从县名查询县编码（如：武义县 → 330723000000000）
- townName: 保持原值（支持多个乡镇，如：王宅镇、白姆乡）
- townCode: 已查询所有乡镇编码（用顿号连接，如：330723105000000、330723201000000）
- riverAlias: 已转换为短简写（如：钱塘江流域 → QTJ）
- subCode: 已转换为拼音简写（如：熟溪 → SX）
- pop_pm: 已自动计算（保护人口/面积）
- gdp_pm: 已自动计算（GDP/面积）
- lake_code, dot_code, geom: 已强制留空

七、详细记录列表
{'-' * 80}
"""
    
    for idx, row in data_df.iterrows():
        report += f"""
{idx + 1}. {row['name']} [{row['code']}]
   - 所属区域: {row['adcdName']}
   - 乡镇: {row['townName']}
   - 面积: {row['area']:.2f} km²
   - 保护人口: {row['proPeople']:.2f} 万人
   - GDP: {row['gdp']:.2f} 万元
   - 防御河流: {row['river']} (流域: {row['riverAlias']})
   - 达标情况: {'达标' if row['isComplete'] == 1 else '不达标'}
"""
    
    report += f"""
八、字段完整性检查
{'-' * 80}
"""
    
    fields_to_check = ['code', 'name', 'adcdName', 'adcdCode', 'townName', 'area',
                       'proPeople', 'gdp', 'pop_pm', 'gdp_pm', 'houseArea', 'roadLen',
                       'protectLand', 'type', 'planType', 'isComplete', 'river', 'proObj',
                       'riverAlias', 'subCode', 'lng', 'lat']
    
    for field in fields_to_check:
        if field in data_df.columns:
            null_count = data_df[field].isnull().sum()
            filled = len(data_df) - null_count
            fill_rate = (filled / len(data_df)) * 100 if len(data_df) > 0 else 0
            status = "✓" if fill_rate == 100 else "○"
            report += f"{status} {field:20s}: {filled:2d}/{len(data_df):2d} ({fill_rate:6.2f}%)\n"
    
    report += f"""
{'=' * 80}
注意事项:
- 所有编码已规范化为大写格式
- adcdName 已从乡镇查询县名
- adcdCode 已从县名查询县编码
- townName 支持多个乡镇（用顿号连接）
- townCode 已查询所有乡镇编码（用顿号连接）
- riverAlias 已转换为短简写（QTJ, OJ等）
- subCode 已转换为河流拼音简写（SX, WX等）
- pop_pm 和 gdp_pm 已自动计算
- lake_code, dot_code, geom 已强制留空
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '保护片数据填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条保护片记录")
    
    river_stats = data_df['river'].value_counts()
    for river_name, count in river_stats.items():
        print(f"✓ {river_name}: {count} 条")
    
    print(f"✓ 达标: {len(data_df[data_df['isComplete'] == 1])} 条, 不达标: {len(data_df[data_df['isComplete'] == 0])} 条")
    print("=" * 80)


def main():
    """主函数"""
    print("\n" + "=" * 80)
    print("保护片数据导入工具 - 熟溪 GeoJSON to Excel (v1.0)")
    print("参考白溪模板，填充熟溪数据")
    print("=" * 80 + "\n")
    
    # 文件路径配置
    geojson_path = 'input/保护片/env.geojson'
    town_csv_path = 'input/city_county_town.csv'
    region_csv_path = 'input/region_name_code.csv'
    excel_path = 'output/datebase_sx.xlsx'  # 目标：熟溪数据库
    sheet_name = '保护片'
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    try:
        # 1. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 2. 加载映射表（乡镇→县，县→编码，乡镇→编码）
        town_to_county = {}
        county_to_code = {}
        town_to_code = {}
        if os.path.exists(town_csv_path) and os.path.exists(region_csv_path):
            town_to_county, county_to_code, town_to_code = load_mappings(town_csv_path, region_csv_path)
        else:
            print(f"⚠ 警告: 映射文件缺失，将无法查询县名和县编码")
        
        # 3. 创建备份
        backup_path = create_backup(excel_path)
        
        # 4. 填充数据
        filled_count = fill_baohuplan_data(geojson_data, excel_path, town_to_county, county_to_code, town_to_code, sheet_name)
        
        # 5. 生成报告
        generate_report(excel_path, geojson_path, filled_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()

