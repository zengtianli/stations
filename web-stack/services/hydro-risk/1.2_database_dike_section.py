#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
堤段数据导入脚本 - 从 GeoJSON 导入到 Excel 数据库

【使用方法】
    python3 fill_dd_to_database.py

【功能】
    1. 从 input/geojson/dd.geojson 读取熟溪堤段数据
    2. 从 input/region_name_code.csv 读取乡镇编码
    3. 参考 database_bx.xlsx（白溪模板）的字段结构
    4. 自动转换拼音简写（钱塘江→QTJ，熟溪→SX等）
    5. 统一编码为大写格式（sxdf0001→SXDF0001）
    6. 填充到 output/datebase_sx.xlsx 的"堤段" sheet
    7. 自动备份原文件
    8. 生成详细报告

【文件要求】
    - input/geojson/dd.geojson       源数据（熟溪堤段GeoJSON）
    - input/region_name_code.csv      乡镇编码对照表
    - templates/database_bx.xlsx      白溪模板（参考字段结构）
    - output/datebase_sx.xlsx         目标Excel（熟溪，必须有"堤段"sheet）

【依赖安装】
    pip install pandas openpyxl

【自定义映射】
    修改脚本第46-70行的映射表：
    RIVER_NAME_TO_CODE    河流名称→编码
    BASIN_NAME_TO_CODE    流域名称→编码
    RIVER_TO_BASIN        河流→流域映射
"""

import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    get_river_code,
    get_basin_code,
    get_basin_name,
    normalize_code,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def load_town_code_mapping(region_csv_path):
    """
    从 CSV 加载乡镇编码映射
    返回: {乡镇编码: 乡镇名称}
    """
    code_to_town = {}
    
    try:
        df_region = pd.read_csv(region_csv_path, encoding='utf-8', dtype={'region_code': str})
        if 'region_name' in df_region.columns and 'region_code' in df_region.columns:
            for _, row in df_region.iterrows():
                region_name = row.get('region_name', '')
                region_code = row.get('region_code', '')
                if region_name and region_code:
                    # 清理编码格式
                    try:
                        code_int = int(float(region_code))
                        code_str = str(code_int)
                        
                        if len(code_str) == 15:
                            # 乡镇级（倒数第4-6位不全是0）
                            if not code_str.endswith('000000000'):
                                code_to_town[code_str] = region_name
                    except:
                        pass
        print(f"✓ 成功加载 {len(code_to_town)} 个乡镇编码→名称映射")
    except Exception as e:
        print(f"⚠ 警告: 无法加载乡镇编码映射 - {e}")
    
    return code_to_town


def fill_dike_section_data(geojson_data, excel_path, code_to_town, sheet_name='堤段'):
    """
    将 GeoJSON 数据填充到 Excel 的堤段 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        code_to_town: 乡镇编码→名称映射
        sheet_name: sheet 名称
    """
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'triCode': 1,           # 流域编码
        'triName': 2,           # 流域名称
        'dikeCode': 3,          # 堤防编码
        'dikeName': 4,          # 堤防名称
        'dsCode': 5,            # 堤段编码
        'dsName': 6,            # 堤段名称
        'river_code': 7,        # 河流编码
        'river_name': 8,        # 河流名称
        'startName': 9,         # 起点
        'endName': 10,          # 终点
        'zya': 11,              # 岸别
        'polderCode': 12,       # 保护片编码
        'ddgc': 13,             # 堤顶高程
        'lc': 14,               # 里程
        'dsLength': 15,         # 堤段长度
        'virtualFl': 16,        # 是否虚拟堤防
        'regionCode': 17,       # 行政区划编码
        'regionName': 18,       # 行政区划名称
        'calRiskL': 19,         # 计算风险等级
        'drz': 20,              # 设计水位
        'grz': 21,              # 警戒水位
        'wrz': 22,              # 保证水位
        'lgtd': 23,             # 经度
        'lttd': 24,             # 纬度
        'geom': 25              # 堤段空间数据
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据
    start_row = 3
    features = geojson_data.get('features', [])
    
    print(f"✓ 开始填充数据（自动处理拼音简写）...")
    
    # 统计转换信息
    conversions = {
        'dikeCode': 0,
        'polderCode': 0,
        'triCode': 0,
    }
    
    for idx, feature in enumerate(features):
        props = feature.get('properties', {})
        row_num = start_row + idx
        
        # 河流信息
        river_name = props.get('river_name', '')
        river_code = props.get('river_code', '')
        
        # 流域信息 - 从河流名称推导
        basin_name = get_basin_name(river_name)
        tri_code = get_basin_code(basin_name)
        if tri_code:
            conversions['triCode'] += 1
        
        ws.cell(row=row_num, column=column_mapping['triCode']).value = tri_code
        ws.cell(row=row_num, column=column_mapping['triName']).value = basin_name
        
        # 堤防编码 - 规范化为大写
        dike_code = normalize_code(props.get('dike_code', ''))
        if dike_code != props.get('dike_code', ''):
            conversions['dikeCode'] += 1
        ws.cell(row=row_num, column=column_mapping['dikeCode']).value = dike_code
        
        # 堤防名称
        ws.cell(row=row_num, column=column_mapping['dikeName']).value = props.get('dike_name', '')
        
        # 堤段信息
        ws.cell(row=row_num, column=column_mapping['dsCode']).value = props.get('ds_code', '')
        ws.cell(row=row_num, column=column_mapping['dsName']).value = props.get('ds_name', '')
        
        # 河流信息
        ws.cell(row=row_num, column=column_mapping['river_code']).value = river_code
        ws.cell(row=row_num, column=column_mapping['river_name']).value = river_name
        
        # 起点终点
        ws.cell(row=row_num, column=column_mapping['startName']).value = props.get('start_name', '')
        ws.cell(row=row_num, column=column_mapping['endName']).value = props.get('end_name', '')
        
        # 岸别
        ws.cell(row=row_num, column=column_mapping['zya']).value = props.get('zya', '')
        
        # 保护片编码 - 规范化为大写
        polder_code = normalize_code(props.get('polder_id', ''))
        if polder_code != props.get('polder_id', ''):
            conversions['polderCode'] += 1
        ws.cell(row=row_num, column=column_mapping['polderCode']).value = polder_code
        
        # 高程和里程
        ddgc = props.get('ddgc')
        if ddgc is not None:
            ws.cell(row=row_num, column=column_mapping['ddgc']).value = round(float(ddgc), 2)
        
        lc = props.get('LC')
        if lc is not None:
            ws.cell(row=row_num, column=column_mapping['lc']).value = round(float(lc), 2)
        
        ds_length = props.get('ds_length')
        if ds_length is not None:
            ws.cell(row=row_num, column=column_mapping['dsLength']).value = round(float(ds_length), 2)
        
        # 虚拟堤防标识（默认0）
        ws.cell(row=row_num, column=column_mapping['virtualFl']).value = props.get('virtualFl', 0)
        
        # 行政区划
        region_code = props.get('regioncode', '')
        ws.cell(row=row_num, column=column_mapping['regionCode']).value = region_code
        
        # 行政区划名称 - 从编码查询或使用原有字段
        region_name = props.get('所属镇', '')
        if not region_name and region_code in code_to_town:
            region_name = code_to_town[region_code]
        ws.cell(row=row_num, column=column_mapping['regionName']).value = region_name
        
        # 风险等级（留空）
        ws.cell(row=row_num, column=column_mapping['calRiskL']).value = ''
        
        # 水位信息
        ws.cell(row=row_num, column=column_mapping['drz']).value = props.get('drz', '')
        ws.cell(row=row_num, column=column_mapping['grz']).value = props.get('grz', '')
        ws.cell(row=row_num, column=column_mapping['wrz']).value = props.get('wrz', '')
        
        # 坐标信息
        lgtd = props.get('lgtd')
        if lgtd is not None:
            ws.cell(row=row_num, column=column_mapping['lgtd']).value = round(float(lgtd), 6)
        
        lttd = props.get('lttd')
        if lttd is not None:
            ws.cell(row=row_num, column=column_mapping['lttd']).value = round(float(lttd), 6)
        
        # 几何信息 - 强制留空
        ws.cell(row=row_num, column=column_mapping['geom']).value = ''
        
        # 显示进度
        if (idx + 1) % 50 == 0 or idx == len(features) - 1:
            print(f"  已填充 {idx + 1}/{len(features)} 条记录...")
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示转换统计
    print(f"\n✓ 编码转换统计:")
    print(f"  - triCode 生成: {conversions['triCode']} 条（从河流名称推导流域）")
    print(f"  - dikeCode 规范化: {conversions['dikeCode']} 条（小写 → 大写）")
    print(f"  - polderCode 规范化: {conversions['polderCode']} 条（小写 → 大写）")
    
    return len(features)


def generate_report(excel_path, geojson_path, filled_count, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='堤段')
    data_df = df.iloc[1:].reset_index(drop=True)  # 跳过表头说明行
    
    report = f"""
{'=' * 80}
堤段数据填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

一、填充统计
{'-' * 80}
总记录数: {filled_count} 条

二、河流分布
{'-' * 80}
{data_df['river_name'].value_counts().to_string()}

三、岸别分布
{'-' * 80}
左岸(L): {len(data_df[data_df['zya'] == 'L'])} 条
右岸(R): {len(data_df[data_df['zya'] == 'R'])} 条

四、行政区分布
{'-' * 80}
{data_df['regionName'].value_counts().to_string()}

五、编码规范化说明
{'-' * 80}
- triCode: 已从河流名称推导流域编码（如：熟溪 → 钱塘江 → QTJ）
- dikeCode: 已规范化为大写（如：sxdf0001 → SXDF0001）
- polderCode: 已规范化为大写（如：sx0001 → SX0001）
- geom: 已强制留空

{'=' * 80}
注意事项:
- 所有编码已规范化为大写格式
- triCode 和 triName 已自动推导
- regionName 已从编码查询或使用原有字段
- geom 已强制留空
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '堤段数据填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条堤段记录")
    
    river_stats = data_df['river_name'].value_counts()
    for river_name, count in river_stats.items():
        print(f"✓ {river_name}: {count} 条")
    
    print(f"✓ 左岸: {len(data_df[data_df['zya'] == 'L'])} 条, 右岸: {len(data_df[data_df['zya'] == 'R'])} 条")
    print("=" * 80)


def main():
    """主函数"""
    print("\n" + "=" * 80)
    print("堤段数据导入工具 - 熟溪 GeoJSON to Excel (v1.0)")
    print("参考白溪模板，填充熟溪数据")
    print("=" * 80 + "\n")
    
    # 文件路径配置
    geojson_path = 'input/geojson/dd.geojson'
    region_csv_path = 'input/region_name_code.csv'
    excel_path = 'output/datebase_sx.xlsx'  # 目标：熟溪数据库
    sheet_name = '堤段'
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    try:
        # 1. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 2. 加载乡镇编码映射
        code_to_town = {}
        if os.path.exists(region_csv_path):
            code_to_town = load_town_code_mapping(region_csv_path)
        else:
            print(f"⚠ 警告: 未找到乡镇编码文件 {region_csv_path}")
        
        # 3. 创建备份
        backup_path = create_backup(excel_path)
        
        # 4. 填充数据
        filled_count = fill_dike_section_data(geojson_data, excel_path, code_to_town, sheet_name)
        
        # 5. 生成报告
        generate_report(excel_path, geojson_path, filled_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()


