#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
堤防数据导入脚本 - 从 GeoJSON 导入到 Excel 数据库

【使用方法】
    python3 fill_df_to_database.py

【功能】
    1. 从 input/df.geojson 读取堤防数据
    2. 自动转换拼音简写（钱塘江流域→QTJLY，熟溪→SX等）
    3. 统一编码为大写格式（sxdf0001→SXDF0001）
    4. 填充到 output/datebase_sx.xlsx 的"堤防" sheet
    5. 自动备份原文件
    6. 生成详细报告

【文件要求】
    - input/df.geojson          源数据（堤防GeoJSON）
    - output/datebase_sx.xlsx   目标Excel（必须有"堤防"sheet）

【依赖安装】
    pip install pandas openpyxl

【自定义映射】
    修改脚本第29-52行的映射表：
    RIVER_NAME_TO_CODE    河流名称→编码
    BASIN_NAME_TO_CODE    流域名称→编码
"""

import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path
import re

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    BASIN_NAME_TO_CODE_LONG,
    get_river_code,
    normalize_code,
    generate_dike_code,
    extract_dike_number,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def get_tri_code(tri_code_name):
    """
    获取流域编码（长版本）
    
    Args:
        tri_code_name: 流域名称
        
    Returns:
        str: 流域编码简写
    """
    if not tri_code_name:
        return ''
    
    # 如果已经是简写，直接返回
    if len(tri_code_name) <= 10 and tri_code_name.isupper():
        return tri_code_name
    
    # 从映射表获取
    return BASIN_NAME_TO_CODE_LONG.get(tri_code_name, tri_code_name)


# ============================================================================
# 核心函数
# ============================================================================

def fill_dike_data(geojson_data, excel_path, sheet_name='堤防'):
    """
    将 GeoJSON 数据填充到 Excel 的堤防 sheet
    
    Args:
        geojson_data: GeoJSON 数据字典
        excel_path: Excel 文件路径
        sheet_name: 要填充的 sheet 名称
        
    Returns:
        int: 填充的记录数
    """
    # 加载 Excel 工作簿
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'triCode': 1,      # 流域编码
        'subCode': 2,      # 支流编码
        'dikeCode': 3,     # 堤防编码
        'dikeName': 4,     # 堤防名称
        'startName': 5,    # 起点名称
        'endName': 6,      # 终点名称
        'zya': 7,          # 左右岸
        'polderCode': 8,   # 圩区编码
        'qdgc': 9,         # 起点高程
        'zdgc': 10,        # 终点高程
        'dsLength': 11,    # 堤防长度
        'lgtd': 12,        # 经度
        'lttd': 13,        # 纬度
        'geom': 14,        # 几何信息
        'type': 15,        # 类型
        'adcdName': 16,    # 行政区名称
        'adcdCode': 17,    # 行政区编码
        'level': 18,       # 等级
        'isGd': 19,        # 是否国道
        'riverName': 20,   # 河流名称
        'river_code': 21,  # 河流编码
        'isComplete': 22,  # 是否完成
        'progress': 23,    # 进度
        'qdlc': 24,        # 起点里程
        'zdlc': 25         # 终点里程
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据（第1行是表头，第2行是说明）
    start_row = 3
    features = geojson_data.get('features', [])
    
    print(f"✓ 开始填充数据（自动处理拼音简写）...")
    
    # 统计转换信息
    conversions = {
        'triCode': 0,
        'river_code': 0,
        'dikeCode': 0,
    }
    
    for idx, feature in enumerate(features):
        props = feature.get('properties', {})
        geom = feature.get('geometry', {})
        row_num = start_row + idx
        
        # 处理 triCode（流域编码）- 转换为拼音简写
        tri_code = props.get('triCode', '')
        tri_code_converted = get_tri_code(tri_code)
        if tri_code != tri_code_converted:
            conversions['triCode'] += 1
        ws.cell(row=row_num, column=column_mapping['triCode']).value = tri_code_converted
        
        # 处理 subCode（支流编码）- 规范化为大写
        sub_code = normalize_code(props.get('subCode', ''), to_upper=True)
        ws.cell(row=row_num, column=column_mapping['subCode']).value = sub_code
        
        # 处理 dikeCode（堤防编码）- 规范化为大写
        dike_code = normalize_code(props.get('dikeCode', ''), to_upper=True)
        if dike_code != props.get('dikeCode', ''):
            conversions['dikeCode'] += 1
        ws.cell(row=row_num, column=column_mapping['dikeCode']).value = dike_code
        
        # 处理 polderCode（圩区编码）- 规范化为大写
        polder_code = normalize_code(props.get('polderCode', ''), to_upper=True)
        ws.cell(row=row_num, column=column_mapping['polderCode']).value = polder_code
        
        # 填充基本信息
        ws.cell(row=row_num, column=column_mapping['dikeName']).value = props.get('dikeName', '')
        ws.cell(row=row_num, column=column_mapping['startName']).value = props.get('startName', '')
        ws.cell(row=row_num, column=column_mapping['endName']).value = props.get('endName', '')
        ws.cell(row=row_num, column=column_mapping['zya']).value = props.get('zya', '')
        ws.cell(row=row_num, column=column_mapping['qdgc']).value = props.get('qdgc', '')
        ws.cell(row=row_num, column=column_mapping['zdgc']).value = props.get('zdgc', '')
        
        # 填充数值类型（保留适当精度）
        ds_length = props.get('dsLength')
        if ds_length is not None:
            ws.cell(row=row_num, column=column_mapping['dsLength']).value = round(float(ds_length), 2)
        
        lgtd = props.get('lgtd')
        if lgtd is not None:
            ws.cell(row=row_num, column=column_mapping['lgtd']).value = round(float(lgtd), 6)
        
        lttd = props.get('lttd')
        if lttd is not None:
            ws.cell(row=row_num, column=column_mapping['lttd']).value = round(float(lttd), 6)
        
        # 填充几何信息（转换为 JSON 字符串）
        if geom:
            ws.cell(row=row_num, column=column_mapping['geom']).value = json.dumps(geom, ensure_ascii=False)
        
        # 填充其他属性
        ws.cell(row=row_num, column=column_mapping['type']).value = props.get('type', '')
        ws.cell(row=row_num, column=column_mapping['adcdName']).value = props.get('adcdName', '')
        ws.cell(row=row_num, column=column_mapping['adcdCode']).value = props.get('adcdCode', '')
        ws.cell(row=row_num, column=column_mapping['level']).value = props.get('level', '')
        ws.cell(row=row_num, column=column_mapping['isGd']).value = props.get('isGd', '')
        
        # 河流名称
        river_name = props.get('riverName', '')
        ws.cell(row=row_num, column=column_mapping['riverName']).value = river_name
        
        # 河流编码 - 自动根据河流名称生成拼音简写
        river_code = get_river_code(river_name)
        if river_code:
            conversions['river_code'] += 1
        ws.cell(row=row_num, column=column_mapping['river_code']).value = river_code
        
        ws.cell(row=row_num, column=column_mapping['isComplete']).value = props.get('isComplete', '')
        ws.cell(row=row_num, column=column_mapping['progress']).value = props.get('progress', '')
        ws.cell(row=row_num, column=column_mapping['qdlc']).value = props.get('qdlc', '')
        ws.cell(row=row_num, column=column_mapping['zdlc']).value = props.get('zdlc', '')
        
        # 显示进度
        if (idx + 1) % 5 == 0 or idx == len(features) - 1:
            print(f"  已填充 {idx + 1}/{len(features)} 条记录...")
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示转换统计
    print(f"\n✓ 编码转换统计:")
    print(f"  - triCode 转换: {conversions['triCode']} 条（流域名称 → 拼音简写）")
    print(f"  - dikeCode 规范化: {conversions['dikeCode']} 条（小写 → 大写）")
    print(f"  - river_code 自动填充: {conversions['river_code']} 条（根据河流名称生成）")
    
    return len(features)


def generate_report(excel_path, geojson_path, filled_count, backup_path):
    """
    生成填充报告
    
    Args:
        excel_path: Excel 文件路径
        geojson_path: GeoJSON 文件路径
        filled_count: 填充的记录数
        backup_path: 备份文件路径
    """
    # 读取填充后的数据
    df = pd.read_excel(excel_path, sheet_name='堤防')
    data_df = df.iloc[1:].reset_index(drop=True)  # 跳过表头说明行
    
    # 生成报告内容
    report = f"""
{'=' * 80}
堤防数据填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

一、填充统计
{'-' * 80}
总记录数: {filled_count} 条

二、河流分布
{'-' * 80}
"""
    
    # 河流统计（同时显示中文名和编码）
    river_stats = data_df.groupby(['riverName', 'river_code']).size()
    for (river_name, river_code), count in river_stats.items():
        report += f"{river_name} ({river_code}): {count} 条\n"
    
    report += f"""
三、左右岸分布
{'-' * 80}
左岸(L): {len(data_df[data_df['zya'] == 'L'])} 条
右岸(R): {len(data_df[data_df['zya'] == 'R'])} 条

四、行政区分布
{'-' * 80}
{data_df['adcdName'].value_counts().to_string()}

五、堤防等级分布
{'-' * 80}
{data_df['level'].value_counts().to_string()}

六、堤防类型分布
{'-' * 80}
{data_df['type'].value_counts().to_string()}

七、编码规范化说明
{'-' * 80}
- triCode: 已转换为拼音简写（如：钱塘江流域 → QTJLY）
- subCode: 已规范化为大写（如：sx → SX）
- dikeCode: 已规范化为大写（如：sxdf0001 → SXDF0001）
- polderCode: 已规范化为大写（如：sx0001 → SX0001）
- river_code: 已自动填充（如：熟溪 → SX）

八、详细记录列表
{'-' * 80}
"""
    
    for idx, row in data_df.iterrows():
        report += f"""
{idx + 1}. {row['dikeName']} [{row['dikeCode']}]
   - 流域编码: {row['triCode']}
   - 支流编码: {row['subCode']}
   - 河流: {row['riverName']} (编码: {row['river_code']})
   - 起点: {row['startName']} → 终点: {row['endName']}
   - 左右岸: {'左岸' if row['zya'] == 'L' else '右岸'}
   - 长度: {row['dsLength']:.2f} 米
   - 行政区: {row['adcdName']}
   - 等级: {row['level']}
"""
    
    report += f"""
九、字段完整性检查
{'-' * 80}
"""
    
    # 检查字段完整性
    fields_to_check = ['triCode', 'subCode', 'dikeCode', 'dikeName', 'startName', 'endName',
                       'zya', 'polderCode', 'dsLength', 'lgtd', 'lttd', 'type',
                       'adcdName', 'adcdCode', 'level', 'isGd', 'riverName', 'river_code', 'isComplete']
    
    for field in fields_to_check:
        null_count = data_df[field].isnull().sum()
        filled = len(data_df) - null_count
        fill_rate = (filled / len(data_df)) * 100 if len(data_df) > 0 else 0
        status = "✓" if fill_rate == 100 else "○"
        report += f"{status} {field:15s}: {filled:2d}/{len(data_df):2d} ({fill_rate:6.2f}%)\n"
    
    report += f"""
{'=' * 80}
注意事项:
- 所有编码已规范化为大写格式
- river_code 已自动根据河流名称填充
- geom 字段包含完整的几何坐标信息（JSON格式）
- qdgc、zdgc、qdlc、zdlc 等字段保持源数据状态
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '堤防数据填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条堤防记录")
    
    # 显示河流统计（带编码）
    for (river_name, river_code), count in river_stats.items():
        print(f"✓ {river_name} ({river_code}): {count} 条")
    
    print(f"✓ 左岸: {len(data_df[data_df['zya'] == 'L'])} 条, 右岸: {len(data_df[data_df['zya'] == 'R'])} 条")
    print("=" * 80)


def main():
    """主函数"""
    print("\n" + "=" * 80)
    print("堤防数据导入工具 - GeoJSON to Excel (v2.0)")
    print("支持拼音简写自动转换")
    print("=" * 80 + "\n")
    
    # 文件路径配置
    geojson_path = 'input/df.geojson'
    excel_path = 'output/datebase_sx.xlsx'
    sheet_name = '堤防'
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    try:
        # 1. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 2. 创建备份
        backup_path = create_backup(excel_path)
        
        # 3. 填充数据
        filled_count = fill_dike_data(geojson_data, excel_path, sheet_name)
        
        # 4. 生成报告
        generate_report(excel_path, geojson_path, filled_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
