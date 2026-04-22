#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
河道中心线数据导入脚本 - 从 GeoJSON 导入到 Excel 数据库

【使用方法】
    python3 fill_hdzxx_to_database.py

【功能】
    1. 从 input/river_center_points.geojson 读取河道中心线数据
    2. 参考 database_bx.xlsx（白溪模板）的字段结构
    3. 自动转换拼音简写（钱塘江→QTJ，熟溪→SX等）
    4. 统一编码为大写格式
    5. 填充到 output/datebase_sx.xlsx 的"河道中心线" sheet
    6. 自动备份原文件
    7. 生成详细报告

【文件要求】
    - input/river_center_points.geojson   源数据（河道中心线GeoJSON）
    - templates/database_bx.xlsx           白溪模板（参考字段结构）
    - output/datebase_sx.xlsx              目标Excel（熟溪，必须有"河道中心线"sheet）

【依赖安装】
    pip install pandas openpyxl

【自定义映射】
    修改脚本第44-70行的映射表：
    RIVER_NAME_TO_CODE    河流名称→编码
    BASIN_NAME_TO_CODE    流域名称→编码
    RIVER_TO_BASIN        河流→流域映射
"""

import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    get_river_code,
    get_basin_code,
    get_basin_name,
    normalize_code,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def fill_river_centerline_data(geojson_data, excel_path, sheet_name='河道中心线'):
    """
    将 GeoJSON 数据填充到 Excel 的河道中心线 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        sheet_name: sheet 名称
    """
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'triCode': 1,           # 流域（干流）编码
        'subCode': 2,           # 支流编码
        'river_code': 3,        # 河道编码
        'river_name': 4,        # 河流名
        'startStName': 5,       # 起始站
        'endStName': 6,         # 终点站
        'lc': 7,                # 里程
        'leftEle': 8,           # 左岸高程
        'rightEle': 9,          # 右岸高程
        'riverBedEle': 10,      # 河底高程
        'leftWrz': 11,          # 左岸警戒水位
        'leftGrz': 12,          # 左岸保证水位
        'leftSjz': 13,          # 左岸设计水位
        'rightWrz': 14,         # 右岸警戒水位
        'rightGrz': 15,         # 右岸保证水位
        'rightSjz': 16,         # 右岸设计水位
        'river_plotline': 17    # 是否河道主线
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据
    start_row = 3
    features = geojson_data.get('features', [])
    
    print(f"✓ 开始填充数据（自动处理拼音简写）...")
    
    # 统计转换信息
    conversions = {
        'triCode': 0,
        'subCode': 0,
        'river_code': 0,
    }
    
    river_stats = {}
    
    for idx, feature in enumerate(features):
        props = feature.get('properties', {})
        row_num = start_row + idx
        
        # 河流信息
        river_name = props.get('NAME', '')
        
        # 流域信息 - 从河流名称推导
        basin_name = get_basin_name(river_name)
        tri_code = get_basin_code(basin_name)
        if tri_code:
            conversions['triCode'] += 1
        
        ws.cell(row=row_num, column=column_mapping['triCode']).value = tri_code
        
        # 支流编码 - 从河流名称转换
        sub_code = get_river_code(river_name)
        sub_code = normalize_code(sub_code)
        if sub_code:
            conversions['subCode'] += 1
        ws.cell(row=row_num, column=column_mapping['subCode']).value = sub_code
        
        # 河道编码 - 与支流编码相同
        river_code = sub_code
        if river_code:
            conversions['river_code'] += 1
        ws.cell(row=row_num, column=column_mapping['river_code']).value = river_code
        
        # 河流名
        ws.cell(row=row_num, column=column_mapping['river_name']).value = river_name
        
        # 起始站和终点站（留空）
        ws.cell(row=row_num, column=column_mapping['startStName']).value = ''
        ws.cell(row=row_num, column=column_mapping['endStName']).value = ''
        
        # 里程
        lc = props.get('LC')
        if lc is not None:
            ws.cell(row=row_num, column=column_mapping['lc']).value = round(float(lc), 2)
        
        # 高程数据
        left_ele = props.get('yagc')  # 左岸高程
        if left_ele is not None:
            ws.cell(row=row_num, column=column_mapping['leftEle']).value = round(float(left_ele), 2)
        
        right_ele = props.get('zagc')  # 右岸高程
        if right_ele is not None:
            ws.cell(row=row_num, column=column_mapping['rightEle']).value = round(float(right_ele), 2)
        
        river_bed_ele = props.get('hdgc')  # 河底高程
        if river_bed_ele is not None:
            ws.cell(row=row_num, column=column_mapping['riverBedEle']).value = round(float(river_bed_ele), 2)
        
        # 水位信息（留空）
        ws.cell(row=row_num, column=column_mapping['leftWrz']).value = ''
        ws.cell(row=row_num, column=column_mapping['leftGrz']).value = ''
        ws.cell(row=row_num, column=column_mapping['leftSjz']).value = ''
        ws.cell(row=row_num, column=column_mapping['rightWrz']).value = ''
        ws.cell(row=row_num, column=column_mapping['rightGrz']).value = ''
        ws.cell(row=row_num, column=column_mapping['rightSjz']).value = ''
        
        # 是否河道主线（默认"是"）
        ws.cell(row=row_num, column=column_mapping['river_plotline']).value = '是'
        
        # 统计河流分布
        river_stats[river_name] = river_stats.get(river_name, 0) + 1
        
        # 显示进度
        if (idx + 1) % 100 == 0 or idx == len(features) - 1:
            print(f"  已填充 {idx + 1}/{len(features)} 条记录...")
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示转换统计
    print(f"\n✓ 编码转换统计:")
    print(f"  - triCode 生成: {conversions['triCode']} 条（从河流名称推导流域）")
    print(f"  - subCode 生成: {conversions['subCode']} 条（河流名称 → 拼音简写）")
    print(f"  - river_code 生成: {conversions['river_code']} 条（同 subCode）")
    
    print(f"\n✓ 河流分布:")
    for river, count in sorted(river_stats.items()):
        code = get_river_code(river)
        print(f"  - {river} ({code}): {count} 条")
    
    return len(features), river_stats


def generate_report(excel_path, geojson_path, filled_count, river_stats, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='河道中心线')
    data_df = df.iloc[1:].reset_index(drop=True)  # 跳过表头说明行
    
    report = f"""
{'=' * 80}
河道中心线数据填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

一、填充统计
{'-' * 80}
总记录数: {filled_count} 条

二、河流分布
{'-' * 80}
"""
    
    for river, count in sorted(river_stats.items()):
        code = get_river_code(river)
        report += f"{river} ({code}): {count} 条\n"
    
    report += f"""
三、里程范围
{'-' * 80}
"""
    
    for river in sorted(river_stats.keys()):
        river_data = data_df[data_df['river_name'] == river]
        if len(river_data) > 0:
            min_lc = river_data['lc'].min()
            max_lc = river_data['lc'].max()
            report += f"{river}: {min_lc:.2f} ~ {max_lc:.2f}\n"
    
    report += f"""
四、高程信息
{'-' * 80}
左岸高程: {data_df['leftEle'].min():.2f} ~ {data_df['leftEle'].max():.2f}
右岸高程: {data_df['rightEle'].min():.2f} ~ {data_df['rightEle'].max():.2f}
河底高程: {data_df['riverBedEle'].min():.2f} ~ {data_df['riverBedEle'].max():.2f}

五、编码规范化说明
{'-' * 80}
- triCode: 已从河流名称推导流域编码（如：熟溪 → 钱塘江 → QTJ）
- subCode: 已转换为拼音简写大写（如：熟溪 → SX）
- river_code: 同 subCode（如：SX）
- river_plotline: 统一设为"是"
- 水位信息: 已全部留空（待后续补充）
- 起始站/终点站: 已全部留空（待后续补充）

{'=' * 80}
注意事项:
- 所有编码已规范化为大写格式
- triCode 已自动推导
- 所有河道均标记为主线（river_plotline="是"）
- 水位信息待后续补充
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '河道中心线数据填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条河道中心线记录")
    
    for river, count in sorted(river_stats.items()):
        code = get_river_code(river)
        print(f"✓ {river} ({code}): {count} 条")
    
    print("=" * 80)


def main():
    """主函数"""
    print("\n" + "=" * 80)
    print("河道中心线数据导入工具 - 熟溪 GeoJSON to Excel (v1.0)")
    print("参考白溪模板，填充熟溪数据")
    print("=" * 80 + "\n")
    
    # 文件路径配置
    geojson_path = 'input/river_center_points.geojson'
    excel_path = 'output/datebase_sx.xlsx'  # 目标：熟溪数据库
    sheet_name = '河道中心线'
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    try:
        # 1. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 2. 创建备份
        backup_path = create_backup(excel_path)
        
        # 3. 填充数据
        filled_count, river_stats = fill_river_centerline_data(geojson_data, excel_path, sheet_name)
        
        # 4. 生成报告
        generate_report(excel_path, geojson_path, filled_count, river_stats, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()


