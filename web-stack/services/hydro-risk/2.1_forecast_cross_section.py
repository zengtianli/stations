#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
主要断面基础信息导入脚本 - 从 GeoJSON 导入到预报调度 Excel

【使用方法】
    python3 fill_dm_to_forecast.py

【功能】
    1. 从 input/dm_lc_ll.geojson 读取断面数据
    2. 参考 forecast_bx.xlsx（白溪模板）的字段结构
    3. 自动转换拼音简写（熟溪→SX等）
    4. 填充到 output/forecast_sx.xlsx 的"主要断面基础信息" sheet
    5. 自动备份原文件
    6. 生成详细报告

【文件要求】
    - input/dm_lc_ll.geojson          源数据（断面GeoJSON）
    - templates/forecast_bx.xlsx      白溪模板（参考字段结构）
    - output/forecast_sx.xlsx         目标Excel（熟溪，必须有"主要断面基础信息"sheet）

【依赖安装】
    pip install pandas openpyxl

【自定义映射】
    修改脚本第39-46行的映射表：
    RIVER_NAME_TO_CODE    河流名称→编码
"""

import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    get_river_code,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def fill_section_info(geojson_data, excel_path, sheet_name='主要断面基础信息'):
    """
    将 GeoJSON 数据填充到 Excel 的主要断面基础信息 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        sheet_name: sheet 名称
    """
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'river_code': 1,        # 河道编码
        'river_name': 2,        # 河流名
        'dot_code': 3,          # 断面编码
        'dot_name': 4,          # 断面名称
        'dot_type': 5,          # 断面类型
        'lc': 6,                # 里程
        'lgtd': 7,              # 经度
        'lttd': 8               # 纬度
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据
    start_row = 3
    features = geojson_data.get('features', [])
    
    # 按 LC 排序（从小到大）
    features_sorted = sorted(features, key=lambda f: f['properties'].get('LC', 0))
    
    print(f"✓ 开始填充数据...")
    
    # 河流信息（默认为熟溪）
    river_name = '熟溪'
    river_code = get_river_code(river_name)
    
    for idx, feature in enumerate(features_sorted):
        props = feature.get('properties', {})
        row_num = start_row + idx
        
        # 河道编码
        ws.cell(row=row_num, column=column_mapping['river_code']).value = river_code
        
        # 河流名
        ws.cell(row=row_num, column=column_mapping['river_name']).value = river_name
        
        # 断面编码（从 1 开始的序号）
        dot_code = idx + 1
        ws.cell(row=row_num, column=column_mapping['dot_code']).value = dot_code
        
        # 断面名称（从 GeoJSON 的"断面编"字段获取，如 sx1, sx2 等）
        dot_name = props.get('断面编', str(dot_code))  # 如果没有则使用编码
        ws.cell(row=row_num, column=column_mapping['dot_name']).value = dot_name
        
        # 断面类型（留空）
        ws.cell(row=row_num, column=column_mapping['dot_type']).value = ''
        
        # 里程
        lc = props.get('LC')
        if lc is not None:
            ws.cell(row=row_num, column=column_mapping['lc']).value = round(float(lc), 2)
        
        # 经度
        longitude = props.get('longitude')
        if longitude is not None:
            ws.cell(row=row_num, column=column_mapping['lgtd']).value = round(float(longitude), 7)
        
        # 纬度
        latitude = props.get('latitude')
        if latitude is not None:
            ws.cell(row=row_num, column=column_mapping['lttd']).value = round(float(latitude), 7)
        
        # 显示进度
        if (idx + 1) % 10 == 0 or idx == len(features_sorted) - 1:
            print(f"  已填充 {idx + 1}/{len(features_sorted)} 条记录...")
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    return len(features_sorted)


def generate_report(excel_path, geojson_path, filled_count, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='主要断面基础信息', header=1)
    data_df = df.reset_index(drop=True)  # header=1 已经正确设置表头，直接使用数据
    
    # Excel 列名映射（中文）
    col_river_code = '*河道编码'
    col_river_name = '*河流名'
    col_dot_code = '*断面编码'
    col_lc = '里程'
    col_lng = '*经度'
    col_lat = '*维度'
    
    report = f"""
{'=' * 80}
主要断面基础信息填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

一、填充统计
{'-' * 80}
总记录数: {filled_count} 条

二、断面编码范围
{'-' * 80}
断面编码: 1 ~ {filled_count}

三、里程范围
{'-' * 80}
里程 (LC): {data_df[col_lc].min():.2f} ~ {data_df[col_lc].max():.2f}

四、坐标范围
{'-' * 80}
经度: {data_df[col_lng].min():.7f} ~ {data_df[col_lng].max():.7f}
纬度: {data_df[col_lat].min():.7f} ~ {data_df[col_lat].max():.7f}

五、字段说明
{'-' * 80}
- river_code: 河道编码（熟溪 → SX）
- river_name: 河流名称（熟溪）
- dot_code: 断面编码（从 1 开始的连续序号）
- dot_name: 断面名称（与 dot_code 相同）
- dot_type: 断面类型（留空）
- lc: 里程（从 GeoJSON 的 LC 字段）
- lgtd: 经度（从 GeoJSON 的 longitude 字段，保留7位小数）
- lttd: 纬度（从 GeoJSON 的 latitude 字段，保留7位小数）

{'=' * 80}
注意事项:
- 断面按里程从小到大排序
- 断面编码为连续序号（1, 2, 3, ...）
- 断面类型暂时留空
- 坐标精度为7位小数
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '断面数据填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条断面记录")
    print(f"✓ 断面编码: 1 ~ {filled_count}")
    print(f"✓ 里程范围: {data_df[col_lc].min():.0f} ~ {data_df[col_lc].max():.0f}")
    print(f"✓ 河流: {data_df[col_river_name].iloc[0]} ({data_df[col_river_code].iloc[0]})")
    print("=" * 80)


def main():
    """主函数"""
    print("\n" + "=" * 80)
    print("主要断面基础信息导入工具 - 熟溪 GeoJSON to Excel (v1.0)")
    print("参考白溪模板，填充熟溪预报调度数据")
    print("=" * 80 + "\n")
    
    # 文件路径配置
    geojson_path = 'input/dm_lc_ll.geojson'
    excel_path = 'output/forecast_sx.xlsx'  # 目标：熟溪预报调度表
    sheet_name = '主要断面基础信息'
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    try:
        # 1. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 2. 创建备份
        backup_path = create_backup(excel_path)
        
        # 3. 填充数据
        filled_count = fill_section_info(geojson_data, excel_path, sheet_name)
        
        # 4. 生成报告
        generate_report(excel_path, geojson_path, filled_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()

