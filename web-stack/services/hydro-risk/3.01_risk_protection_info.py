#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
保护片信息导入脚本 - 从 GeoJSON 导入到风险分析 Excel

【脚本编号】3.01
【目标文件】风险分析-{流域}.xlsx
【目标Sheet】保护片信息 

【使用方法】
    # 使用默认路径
    python3 3.01_*.py
    
    # 指定路径
    python3 3.01_*.py -g /path/to/data.geojson -e /path/to/risk.xlsx

【功能】
    1. 从 input/保护片/env.geojson 读取保护片数据
    2. 参考模板的字段结构
    3. 自动转换拼音简写和编码大写
    4. 填充到 output/risk_xx.xlsx 的"保护片信息 " sheet
    5. 自动备份原文件
    6. 生成详细报告

【文件要求】
    - input/保护片/env.geojson                  源数据（熟溪保护片GeoJSON）
    - input/GDP、人口、房屋、耕地、道路重新计算结果.csv  最新数据（优先级最高）
    - templates/risk_bx.xlsx                    白溪模板（参考字段结构）
    - output/risk_sx.xlsx                       目标Excel（熟溪，必须有"保护片信息"sheet）

【依赖安装】
    pip install pandas openpyxl

【字段映射】
    - code -> polder_code（保护片编码，大写）
    - name -> polder_name（保护片名称）
    - area -> areas（区域面积，平方公里）
    - protectLand -> cultivated_area（耕地面积，公顷→平方公里，除以100）
    - proPeople -> pop（人口，万人）
    - gdp -> gdp（GDP，万元）
    - planType -> gh_flood（规划防洪标准）
    - type -> xz_flood（现在防洪标准，提取数字）
    - isComplete -> up_standard（是否达标）
"""

import argparse
import pandas as pd
import json
import re
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    normalize_code,
)
from excel_ops import (
    load_latest_data,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def extract_flood_standard(type_str):
    """从类型字符串中提取防洪标准数字（如 "5年一遇" -> 5）"""
    if not type_str:
        return ''
    match = re.search(r'(\d+)年一遇', str(type_str))
    if match:
        return int(match.group(1))
    return ''


def load_latest_gdp_data(csv_path):
    """
    加载最新的GDP、人口、房屋、耕地、道路数据（优先级最高）
    返回: {保护片名称: {字段: 值}}
    """
    latest_data = {}
    
    try:
        df = pd.read_csv(csv_path, encoding='utf-8')
        print(f"✓ 成功读取最新数据文件: {csv_path}")
        
        # 遍历CSV数据
        for _, row in df.iterrows():
            name = row.get('name', '')
            if name:
                latest_data[name] = {
                    'gdp': row.get('GDP(万元）', None),
                    'pop': row.get('pop（万人）', None),
                    'cultivated_area': row.get('Vegetation（公顷）', None),  # 公顷
                    'road': row.get('road（km）', None),
                    'house_area': row.get('房屋面（万㎡）', None)
                }
        
        print(f"  包含 {len(latest_data)} 条最新数据记录（优先级最高）")
        return latest_data
        
    except FileNotFoundError:
        print(f"⚠ 警告: 未找到最新数据文件 {csv_path}，将使用GeoJSON中的数据")
        return {}
    except Exception as e:
        print(f"⚠ 警告: 读取最新数据文件失败 - {e}，将使用GeoJSON中的数据")
        return {}


def fill_polder_info(geojson_data, excel_path, latest_data=None, sheet_name='保护片信息'):
    """
    将 GeoJSON 数据填充到 Excel 的保护片信息 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        latest_data: 最新数据字典（优先级最高）
        sheet_name: sheet 名称
    """
    if latest_data is None:
        latest_data = {}
    
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'polder_code': 1,           # 保护片编码
        'polder_name': 2,           # 保护片名称
        'areas': 3,                 # 区域面积（平方公里）
        'cultivated_area': 4,       # 耕地面积（平方公里）
        'pop': 5,                   # 人口（万人）
        'gdp': 6,                   # GDP（万元）
        'design_level': 7,          # 设计水位
        'flood_standard': 8,        # 防洪标准
        'gh_flood': 9,              # 规划防洪标准
        'xz_flood': 10,             # 现在防洪标准
        'up_standard': 11,          # 是否达标
        'year': 12,                 # 实施年份
        'lgtd': 13,                 # 中心点经度
        'lttd': 14,                 # 中心点纬度
        'ls_lgtd': 15,              # 左上角经度
        'ls_lttd': 16,              # 左上角纬度
        'rx_lgtd': 17,              # 右下角经度
        'rx_lttd': 18,              # 右下角纬度
        'dike_code': 19,            # 关联的堤段编码
        'lake_code': 20,            # 关联的概湖编码
        'dot_code': 21              # 关联的断面编码
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据
    start_row = 3
    features = geojson_data.get('features', [])
    
    print(f"✓ 开始填充数据...")
    
    # 统计转换信息
    conversions = {
        'polder_code': 0,
        'cultivated_area': 0,
        'xz_flood': 0,
    }
    
    # 统计CSV数据使用情况
    csv_usage = {
        'gdp': 0,
        'pop': 0,
        'cultivated_area': 0,
    }
    
    for idx, feature in enumerate(features):
        props = feature.get('properties', {})
        row_num = start_row + idx
        
        # 保护片编码 - 规范化为大写
        polder_code = normalize_code(props.get('code', ''))
        if polder_code != props.get('code', ''):
            conversions['polder_code'] += 1
        ws.cell(row=row_num, column=column_mapping['polder_code']).value = polder_code
        
        # 保护片名称
        polder_name = props.get('name', '')
        ws.cell(row=row_num, column=column_mapping['polder_name']).value = polder_name
        
        # 获取该保护片的最新数据（如果有）
        latest = latest_data.get(polder_name, {})
        
        # 区域面积（平方公里）- 始终使用GeoJSON的数据
        area = props.get('area')
        if area is not None:
            ws.cell(row=row_num, column=column_mapping['areas']).value = round(float(area), 6)
        
        # 耕地面积（公顷 -> 平方公里，除以100）- 优先使用CSV数据
        if latest.get('cultivated_area') is not None:
            # 使用CSV的最新数据
            cultivated_area = float(latest['cultivated_area']) / 100  # 公顷转平方公里
            ws.cell(row=row_num, column=column_mapping['cultivated_area']).value = round(cultivated_area, 6)
            conversions['cultivated_area'] += 1
            csv_usage['cultivated_area'] += 1
        else:
            # 使用GeoJSON数据
            protect_land = props.get('protectLand')
            if protect_land is not None:
                cultivated_area = float(protect_land) / 100  # 公顷转平方公里
                ws.cell(row=row_num, column=column_mapping['cultivated_area']).value = round(cultivated_area, 6)
                conversions['cultivated_area'] += 1
        
        # 人口（万人）- 优先使用CSV数据
        if latest.get('pop') is not None:
            # 使用CSV的最新数据
            ws.cell(row=row_num, column=column_mapping['pop']).value = round(float(latest['pop']), 6)
            csv_usage['pop'] += 1
        else:
            # 使用GeoJSON数据
            pop = props.get('proPeople')
            if pop is not None:
                ws.cell(row=row_num, column=column_mapping['pop']).value = round(float(pop), 4)
        
        # GDP（万元）- 优先使用CSV数据
        if latest.get('gdp') is not None:
            # 使用CSV的最新数据
            ws.cell(row=row_num, column=column_mapping['gdp']).value = round(float(latest['gdp']), 4)
            csv_usage['gdp'] += 1
        else:
            # 使用GeoJSON数据
            gdp = props.get('gdp')
            if gdp is not None:
                ws.cell(row=row_num, column=column_mapping['gdp']).value = round(float(gdp), 4)
        
        # 设计水位（留空）
        ws.cell(row=row_num, column=column_mapping['design_level']).value = ''
        
        # 防洪标准（留空）
        ws.cell(row=row_num, column=column_mapping['flood_standard']).value = ''
        
        # 规划防洪标准
        plan_type = props.get('planType')
        if plan_type is not None:
            ws.cell(row=row_num, column=column_mapping['gh_flood']).value = int(plan_type)
        
        # 现在防洪标准 - 从类型字段提取数字
        type_str = props.get('type', '')
        xz_flood = extract_flood_standard(type_str)
        if xz_flood:
            conversions['xz_flood'] += 1
        ws.cell(row=row_num, column=column_mapping['xz_flood']).value = xz_flood
        
        # 是否达标
        is_complete = props.get('isComplete')
        if is_complete is not None:
            ws.cell(row=row_num, column=column_mapping['up_standard']).value = int(is_complete)
        
        # 实施年份（留空）
        ws.cell(row=row_num, column=column_mapping['year']).value = ''
        
        # 中心点坐标
        lng = props.get('lng')
        if lng is not None:
            ws.cell(row=row_num, column=column_mapping['lgtd']).value = round(float(lng), 6)
        
        lat = props.get('lat')
        if lat is not None:
            ws.cell(row=row_num, column=column_mapping['lttd']).value = round(float(lat), 6)
        
        # 左上角坐标
        ls_lgtd = props.get('topLeftLongitude')
        if ls_lgtd is not None:
            ws.cell(row=row_num, column=column_mapping['ls_lgtd']).value = round(float(ls_lgtd), 6)
        
        ls_lttd = props.get('topLeftLatitude')
        if ls_lttd is not None:
            ws.cell(row=row_num, column=column_mapping['ls_lttd']).value = round(float(ls_lttd), 6)
        
        # 右下角坐标
        rx_lgtd = props.get('bottomRightLongitude')
        if rx_lgtd is not None:
            ws.cell(row=row_num, column=column_mapping['rx_lgtd']).value = round(float(rx_lgtd), 6)
        
        rx_lttd = props.get('bottomRightLatitude')
        if rx_lttd is not None:
            ws.cell(row=row_num, column=column_mapping['rx_lttd']).value = round(float(rx_lttd), 6)
        
        # 关联编码（留空）
        ws.cell(row=row_num, column=column_mapping['dike_code']).value = ''
        ws.cell(row=row_num, column=column_mapping['lake_code']).value = ''
        ws.cell(row=row_num, column=column_mapping['dot_code']).value = ''
        
        # 显示进度
        if (idx + 1) % 5 == 0 or idx == len(features) - 1:
            print(f"  已填充 {idx + 1}/{len(features)} 条记录...")
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示转换统计
    print(f"\n✓ 数据转换统计:")
    print(f"  - polder_code 规范化: {conversions['polder_code']} 条（小写 → 大写）")
    print(f"  - cultivated_area 转换: {conversions['cultivated_area']} 条（公顷 → 平方公里）")
    print(f"  - xz_flood 提取: {conversions['xz_flood']} 条（从'X年一遇'提取数字）")
    
    # 显示CSV数据使用统计
    print(f"\n✓ CSV最新数据使用统计（优先级最高）:")
    print(f"  - GDP: {csv_usage['gdp']}/{len(features)} 条使用CSV数据")
    print(f"  - 人口: {csv_usage['pop']}/{len(features)} 条使用CSV数据")
    print(f"  - 耕地面积: {csv_usage['cultivated_area']}/{len(features)} 条使用CSV数据")
    
    return len(features)


def generate_report(excel_path, geojson_path, filled_count, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='保护片信息 ', header=1)  # 注意：sheet名称末尾有空格
    data_df = df.reset_index(drop=True)
    
    # Excel 列名映射（中文）
    col_polder_code = '*保护片编码'
    col_polder_name = '*保护片名称'
    col_areas = '*区域面积（平方公里）'
    col_pop = '*人口（万人）'
    col_gdp = '*Gdp（万元）'
    
    report = f"""
{'=' * 80}
保护片信息填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

一、填充统计
{'-' * 80}
总记录数: {filled_count} 条

二、保护片列表
{'-' * 80}
"""
    
    for idx, row in data_df.iterrows():
        code = row[col_polder_code]
        name = row[col_polder_name]
        if pd.notna(code):
            report += f"{code}: {name}\n"
    
    report += f"""
三、面积统计
{'-' * 80}
区域面积: {data_df[col_areas].min():.2f} ~ {data_df[col_areas].max():.2f} 平方公里
总面积: {data_df[col_areas].sum():.2f} 平方公里

四、人口与GDP
{'-' * 80}
总人口: {data_df[col_pop].sum():.2f} 万人
总GDP: {data_df[col_gdp].sum():.2f} 万元

五、字段说明
{'-' * 80}
- polder_code: 保护片编码（大写，如 SX0001）
- areas: 区域面积（从 GeoJSON area 字段）
- cultivated_area: 耕地面积（从 protectLand 转换，公顷→平方公里）
- pop: 人口（从 proPeople 字段）
- gdp: GDP（从 gdp 字段）
- gh_flood: 规划防洪标准（从 planType 字段）
- xz_flood: 现在防洪标准（从 type 字段提取数字）
- up_standard: 是否达标（从 isComplete 字段）
- 坐标: 中心点、左上角、右下角坐标均已填充
- 关联编码: dike_code, lake_code, dot_code 留空

{'=' * 80}
注意事项:
- 所有编码已规范化为大写格式
- 耕地面积已从公顷转换为平方公里
- 现在防洪标准已从"X年一遇"提取数字
- 关联编码待后续补充
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '保护片信息填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条保护片记录")
    print(f"✓ 总面积: {data_df[col_areas].sum():.2f} 平方公里")
    print(f"✓ 总人口: {data_df[col_pop].sum():.2f} 万人")
    print(f"✓ 总GDP: {data_df[col_gdp].sum():.2f} 万元")
    print("=" * 80)


def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='保护片信息填充',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例: python3 3.01_risk_protection_info.py -g /path/to/env.geojson -e /path/to/hx.xlsx'
    )
    parser.add_argument('-g', '--geojson', help='GeoJSON 文件路径 (env.geojson)')
    parser.add_argument('-e', '--excel', help='Excel 文件路径 (risk_xx.xlsx)')
    parser.add_argument('-d', '--data', help='CSV 数据文件路径 (可选)')
    parser.add_argument('-s', '--sheet', help='Sheet 名称', default='保护片信息 ')
    args = parser.parse_args()
    
    geojson_path = args.geojson or 'input/保护片/env.geojson'
    excel_path = args.excel or 'output/risk_sx.xlsx'
    csv_path = args.data or 'input/GDP、人口、房屋、耕地、道路重新计算结果.csv'
    sheet_name = args.sheet
    
    print("\n" + "=" * 80)
    print("保护片信息导入工具 (v2.0)")
    print("=" * 80)
    print(f"  GeoJSON: {geojson_path}")
    print(f"  Excel:   {excel_path}")
    print(f"  CSV:     {csv_path}")
    print("=" * 80 + "\n")
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        print(f"  当前工作目录: {os.getcwd()}")
        sys.exit(1)
    
    try:
        # 1. 加载最新数据（优先级最高）
        latest_data = load_latest_gdp_data(csv_path)
        
        # 2. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 3. 创建备份
        backup_path = create_backup(excel_path)
        
        # 4. 填充数据（传入最新数据）
        filled_count = fill_polder_info(geojson_data, excel_path, latest_data, sheet_name)
        
        # 5. 生成报告
        generate_report(excel_path, geojson_path, filled_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()

