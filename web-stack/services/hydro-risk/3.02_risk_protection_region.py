#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
保护片行政区域信息导入脚本 - 从 GeoJSON 导入到风险分析 Excel

【脚本编号】3.02
【目标文件】风险分析-{流域}.xlsx
【目标Sheet】保护片行政区域信息  

【使用方法】
    # 使用默认路径
    python3 3.02_*.py
    
    # 指定路径
    python3 3.02_*.py -g /path/to/data.geojson -e /path/to/risk.xlsx

【功能】
    1. 从 input/保护片/env.geojson 读取保护片数据
    2. 从 input/GDP、人口、房屋、耕地、道路重新计算结果.csv 读取最新数据（优先级最高）
    3. 参考模板的字段结构
    4. 处理跨乡镇的保护片（生成多行记录）
    5. 填充到 output/risk_xx.xlsx 的"保护片行政区域信息  " sheet
    6. 自动备份原文件
    7. 生成详细报告

【文件要求】
    - input/保护片/env.geojson                  源数据（熟溪保护片GeoJSON）
    - input/GDP、人口、房屋、耕地、道路重新计算结果.csv  最新数据（优先级最高）
    - input/region_name_code.csv                乡镇编码对照表
    - templates/risk_bx.xlsx                    白溪模板（参考字段结构）
    - output/risk_sx.xlsx                       目标Excel（熟溪）

【依赖安装】
    pip install pandas openpyxl

【字段说明】
    - polder_code: 保护片编码（大写）
    - region_code: 区域编码（乡镇编码，如果跨多个乡镇会有多行）
    - weight: 权重（如果跨N个乡镇，每个为1/N）
    - pop_pm: 每平方千米人口（总人口/总面积）
    - gdp_pm: 每平方千米GDP（总GDP/总面积）
"""
import argparse
import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    normalize_code,
)
from excel_ops import (
    load_latest_data,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def load_town_codes(region_csv_path):
    """
    加载乡镇编码映射
    返回: {乡镇名称: 编码}
    """
    town_to_code = {}
    
    try:
        df_region = pd.read_csv(region_csv_path, encoding='utf-8', dtype={'region_code': str})
        if 'region_name' in df_region.columns and 'region_code' in df_region.columns:
            for _, row in df_region.iterrows():
                region_name = row.get('region_name', '')
                region_code = row.get('region_code', '')
                if region_name and region_code:
                    try:
                        code_int = int(float(region_code))
                        code_str = str(code_int)
                        
                        if len(code_str) == 15:
                            # 乡镇级（倒数第4-6位不全是0）
                            if not code_str.endswith('000000000'):
                                town_to_code[region_name] = code_str
                    except:
                        pass
        print(f"✓ 成功加载 {len(town_to_code)} 个乡镇编码映射")
    except Exception as e:
        print(f"⚠ 警告: 无法加载乡镇编码映射 - {e}")
    
    return town_to_code


def fill_region_info(geojson_data, excel_path, latest_data, town_to_code, sheet_name='保护片行政区域信息  '):
    """
    将 GeoJSON 数据填充到 Excel 的保护片行政区域信息 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        latest_data: 最新数据字典
        town_to_code: 乡镇编码映射
        sheet_name: sheet 名称
    """
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'polder_code': 1,       # 保护片编码
        'region_code': 2,       # 区域编码（乡镇编码）
        'weight': 3,            # 权重
        'pop_pm': 4,            # 每平方千米人口
        'gdp_pm': 5             # 每平方千米GDP
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据
    start_row = 3
    features = geojson_data.get('features', [])
    
    print(f"✓ 开始填充数据...")
    
    current_row = start_row
    multi_town_count = 0
    total_records = 0
    
    for idx, feature in enumerate(features):
        props = feature.get('properties', {})
        
        # 保护片编码 - 规范化为大写
        polder_code = normalize_code(props.get('code', ''))
        polder_name = props.get('name', '')
        
        # 获取乡镇名称列表（从 adcdName 字段，可能有多个）
        town_names_str = props.get('adcdName', '')
        town_names = [t.strip() for t in town_names_str.split('、') if t.strip()]
        
        # 获取最新数据
        latest = latest_data.get(polder_name, {})
        
        # 获取面积、人口、GDP
        area = props.get('area')  # 平方公里
        
        # 优先使用CSV数据
        if latest.get('pop') is not None:
            pop = float(latest['pop'])  # 万人
        else:
            pop = props.get('proPeople')  # 万人
            if pop is not None:
                pop = float(pop)
        
        if latest.get('gdp') is not None:
            gdp = float(latest['gdp'])  # 万元
        else:
            gdp = props.get('gdp')  # 万元
            if gdp is not None:
                gdp = float(gdp)
        
        # 计算每平方千米人口和GDP
        pop_pm = None
        gdp_pm = None
        if area and area > 0:
            if pop is not None:
                pop_pm = pop / float(area)  # 万人/平方公里
            if gdp is not None:
                gdp_pm = gdp / float(area)  # 万元/平方公里
        
        # 如果没有乡镇信息，跳过
        if not town_names:
            continue
        
        # 计算权重（如果跨多个乡镇，平均分配）
        num_towns = len(town_names)
        weight = 1.0 / num_towns if num_towns > 0 else 1.0
        
        if num_towns > 1:
            multi_town_count += 1
        
        # 为每个乡镇生成一行记录
        for town_name in town_names:
            # 查询乡镇编码
            region_code = town_to_code.get(town_name, '')
            
            if not region_code:
                print(f"  ⚠ 警告: 未找到乡镇'{town_name}'的编码")
                continue
            
            # 填充数据
            ws.cell(row=current_row, column=column_mapping['polder_code']).value = polder_code
            ws.cell(row=current_row, column=column_mapping['region_code']).value = region_code
            ws.cell(row=current_row, column=column_mapping['weight']).value = round(weight, 6)
            
            if pop_pm is not None:
                ws.cell(row=current_row, column=column_mapping['pop_pm']).value = round(pop_pm, 6)
            
            if gdp_pm is not None:
                ws.cell(row=current_row, column=column_mapping['gdp_pm']).value = round(gdp_pm, 6)
            
            current_row += 1
            total_records += 1
        
        # 显示进度
        if (idx + 1) % 5 == 0 or idx == len(features) - 1:
            print(f"  已处理 {idx + 1}/{len(features)} 个保护片...")
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示统计
    print(f"\n✓ 数据统计:")
    print(f"  - 保护片数量: {len(features)} 个")
    print(f"  - 跨乡镇保护片: {multi_town_count} 个")
    print(f"  - 生成记录数: {total_records} 条")
    
    return total_records

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
def generate_report(excel_path, geojson_path, filled_count, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='保护片行政区域信息  ', header=1)
    data_df = df.reset_index(drop=True)
    
    report = f"""
{'=' * 80}
保护片行政区域信息填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

# from xlsx_common import normalize_code, read_geojson, create_backup
一、填充统计
{'-' * 80}
总记录数: {filled_count} 条
唯一保护片数: {data_df['*保护片编码'].nunique()} 个

# from xlsx_common import normalize_code, read_geojson, create_backup
二、权重分布
{'-' * 80}
权重=1（单乡镇）: {len(data_df[data_df['*权重'] == 1])} 条
权重<1（跨乡镇）: {len(data_df[data_df['*权重'] < 1])} 条

# from xlsx_common import normalize_code, read_geojson, create_backup
三、字段说明
{'-' * 80}
- polder_code: 保护片编码（大写）
- region_code: 区域编码（乡镇编码，15位）
- weight: 权重（跨N个乡镇时，每个为1/N）
- pop_pm: 每平方千米人口（万人/平方公里）
- gdp_pm: 每平方千米GDP（万元/平方公里）

# from xlsx_common import normalize_code, read_geojson, create_backup
四、数据来源
{'-' * 80}
- 保护片编码: GeoJSON code字段
- 区域编码: 从region_name_code.csv查询
- 人口/GDP: 优先使用CSV最新数据
- 面积: GeoJSON area字段

# from xlsx_common import normalize_code, read_geojson, create_backup
{'=' * 80}
注意事项:
- 跨乡镇保护片会生成多行记录
- 权重按乡镇数量平均分配
- pop_pm和gdp_pm是整个保护片的指标
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '保护片行政区域信息填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条行政区域记录")
    print(f"✓ 涉及 {data_df['*保护片编码'].nunique()} 个保护片")
    print(f"✓ 涉及 {data_df['*区域编码'].nunique()} 个乡镇")
    print("=" * 80)

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='保护片行政区域信息填充',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例: python3 3.02_risk_protection_region.py -g /path/to/env.geojson -e /path/to/hx.xlsx'
    )
    parser.add_argument('-g', '--geojson', help='GeoJSON 文件路径 (env.geojson)')
    parser.add_argument('-e', '--excel', help='Excel 文件路径 (risk_xx.xlsx)')
    parser.add_argument('-d', '--data', help='CSV 数据文件路径')
    parser.add_argument('-r', '--region', help='乡镇编码对照表路径')
    parser.add_argument('-s', '--sheet', help='Sheet 名称', default='保护片行政区域信息  ')
    args = parser.parse_args()
    
    geojson_path = args.geojson or 'input/保护片/env.geojson'
    csv_path = args.data or 'input/GDP、人口、房屋、耕地、道路重新计算结果.csv'
    region_csv_path = args.region or 'input/region_name_code.csv'
    excel_path = args.excel or 'output/risk_sx.xlsx'
    sheet_name = args.sheet
    
    print("\n" + "=" * 80)
    print("保护片行政区域信息导入工具 (v2.0)")
    print("=" * 80)
    print(f"  GeoJSON: {geojson_path}")
    print(f"  Excel:   {excel_path}")
    print("=" * 80 + "\n")
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        sys.exit(1)
    
    try:
        # 1. 加载最新数据（优先级最高）
        latest_data = load_latest_data(csv_path)
        
        # 2. 加载乡镇编码映射
        town_to_code = load_town_codes(region_csv_path)
        
        # 3. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 4. 创建备份
        backup_path = create_backup(excel_path)
        
        # 5. 填充数据
        filled_count = fill_region_info(geojson_data, excel_path, latest_data, town_to_code, sheet_name)
        
        # 6. 生成报告
        generate_report(excel_path, geojson_path, filled_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
if __name__ == '__main__':
    main()

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
