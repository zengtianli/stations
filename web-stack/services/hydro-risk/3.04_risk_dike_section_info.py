#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
堤段对应信息导入脚本 - 从 GeoJSON 导入到风险分析 Excel

【脚本编号】3.04
【目标文件】风险分析-{流域}.xlsx
【目标Sheet】堤段对应信息

【使用方法】
    # 使用默认路径（熟溪）
    python3 3.04_risk_dike_section_info.py
    
    # 指定路径（任意流域）
    python3 3.04_risk_dike_section_info.py -g /path/to/dd.geojson -e /path/to/risk.xlsx

【功能】
    1. 从 input/geojson/dd_fix.geojson 读取堤段数据
    2. 提取堤段详细信息（河道、堤防、堤段编码、高程、里程、坐标等）
    3. 根据堤顶高程自动计算水位（设计、保证、警戒）
    4. 参考模板的字段结构
    5. 填充到 output/risk_xx.xlsx 的"堤段对应信息" sheet
    6. 自动备份原文件
    7. 生成详细报告

【文件要求】
    - input/geojson/dd_fix.geojson    源数据（熟溪堤段GeoJSON，含river_name）
    - templates/risk_bx.xlsx       白溪模板（参考字段结构）
    - output/risk_sx.xlsx          目标Excel（熟溪）

【依赖安装】
    pip install pandas openpyxl

【字段说明】
    - river_code: 河道编码（大写，如SX）
    - river_name: 河道名称（如熟溪）
    - dike_code: 堤防编码（大写，如SXDF0003）
    - dike_name: 堤防名称（如熟溪堤防0003）
    - ds_code: 堤段编码（如20250001）
    - ds_name: 堤段名称（如20250001）
    - zya: 岸别（L/R，左岸/右岸）
    - polder_code: 保护片编码（大写，如SX0003）
    - ddgc: 堤顶高程（米）
    - lc: 里程（米）
    - ds_length: 堤段长度（米）
    - lgtd: 中心点经度
    - lttd: 中心点纬度
    - region_code: 区域编码（乡镇编码）
    - drz: 设计水位（米，自动计算 = 堤顶高程 - 0.5）
    - grz: 保证水位（米，自动计算 = 堤顶高程 - 1.0）
    - wrz: 警戒水位（米，自动计算 = 堤顶高程 - 1.5）
    
【水位计算规则】
    设计水位 = 堤顶高程 - 0.5
    保证水位 = 堤顶高程 - 1.0
    警戒水位 = 堤顶高程 - 1.5
"""
import pandas as pd
import json
import re
import argparse
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    normalize_code,
    extract_dike_number,
    generate_dike_code,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def fill_dike_section_info(geojson_data, excel_path, sheet_name='堤段对应信息'):
    """
    将 GeoJSON 数据填充到 Excel 的堤段对应信息 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        sheet_name: sheet 名称
    """
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'river_code': 1,      # 河道编码
        'river_name': 2,      # 河道名称
        'dike_code': 3,       # 堤防编码
        'dike_name': 4,       # 堤防名称
        'ds_code': 5,         # 堤段编码
        'ds_name': 6,         # 堤段名称
        'zya': 7,             # 岸别（L/R）
        'polder_code': 8,     # 保护片编码
        'ddgc': 9,            # 堤顶高程
        'lc': 10,             # 里程
        'ds_length': 11,      # 堤段长度
        'lgtd': 12,           # 中心点经度
        'lttd': 13,           # 中心点纬度
        'region_code': 14,    # 区域编码
        'drz': 15,            # 设计水位
        'grz': 16,            # 保证水位
        'wrz': 17             # 警戒水位
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据
    start_row = 3
    features = geojson_data.get('features', [])
    
    print(f"✓ 开始填充数据...")
    
    current_row = start_row
    polder_set = set()  # 用于统计唯一保护片数
    dike_set = set()    # 用于统计唯一堤防数
    
    for idx, feature in enumerate(features):
        props = feature.get('properties', {})
        
        # 河道编码和名称 - 规范化为大写
        river_code = normalize_code(props.get('river_code', ''))
        river_name = props.get('river_name', '')
        
        # 如果 river_name 为空，根据 river_code 推断
        if not river_name:
            river_code_to_name = {
                'SX': '熟溪',
                'WX': '武溪',
                'GZX': '古竹溪',
                'BX': '白溪',
                'HX': '华溪',
            }
            river_name = river_code_to_name.get(river_code.upper(), '')
        
        # 堤防编码和名称 - 直接使用源数据的编码（如 SXDF0003）
        dike_name = props.get('dike_name', '')
        dike_code = normalize_code(props.get('dike_code', ''))
        
        # 堤段编码和名称
        ds_code = props.get('ds_code', '')
        ds_name = props.get('ds_name', '')
        
        # 岸别（L/R）
        zya = props.get('zya', '').upper() if props.get('zya') else ''
        
        # 保护片编码 - 规范化为大写
        polder_id = props.get('polder_id', '')
        polder_code = normalize_code(polder_id)
        
        # 堤顶高程
        ddgc = props.get('ddgc')
        
        # 里程
        lc = props.get('LC')
        
        # 堤段长度
        ds_length = props.get('ds_length')
        
        # 中心点经纬度
        lgtd = props.get('lgtd')
        lttd = props.get('lttd')
        
        # 区域编码
        region_code = props.get('regioncode', '')
        
        # 设计水位、保证水位、警戒水位（根据堤顶高程计算）
        # 设计水位 = 堤顶高程 - 0.5
        # 保证水位 = 堤顶高程 - 1.0
        # 警戒水位 = 堤顶高程 - 1.5
        drz = None
        grz = None
        wrz = None
        if ddgc is not None:
            drz = float(ddgc) - 0.5
            grz = float(ddgc) - 1.0
            wrz = float(ddgc) - 1.5
        
        # 跳过无效记录（至少需要堤段编码）
        if not ds_code:
            continue
        
        # 填充数据
        ws.cell(row=current_row, column=column_mapping['river_code']).value = river_code
        ws.cell(row=current_row, column=column_mapping['river_name']).value = river_name
        ws.cell(row=current_row, column=column_mapping['dike_code']).value = dike_code
        ws.cell(row=current_row, column=column_mapping['dike_name']).value = dike_name
        ws.cell(row=current_row, column=column_mapping['ds_code']).value = str(ds_code)
        ws.cell(row=current_row, column=column_mapping['ds_name']).value = str(ds_name)
        ws.cell(row=current_row, column=column_mapping['zya']).value = zya
        ws.cell(row=current_row, column=column_mapping['polder_code']).value = polder_code
        
        # 数值字段（可能为None）
        if ddgc is not None:
            ws.cell(row=current_row, column=column_mapping['ddgc']).value = float(ddgc)
        
        if lc is not None:
            ws.cell(row=current_row, column=column_mapping['lc']).value = float(lc)
        
        if ds_length is not None:
            ws.cell(row=current_row, column=column_mapping['ds_length']).value = float(ds_length)
        
        if lgtd is not None:
            ws.cell(row=current_row, column=column_mapping['lgtd']).value = float(lgtd)
        
        if lttd is not None:
            ws.cell(row=current_row, column=column_mapping['lttd']).value = float(lttd)
        
        ws.cell(row=current_row, column=column_mapping['region_code']).value = str(region_code) if region_code else ''
        
        # 水位信息（可能为None）
        if drz is not None:
            ws.cell(row=current_row, column=column_mapping['drz']).value = float(drz)
        
        if grz is not None:
            ws.cell(row=current_row, column=column_mapping['grz']).value = float(grz)
        
        if wrz is not None:
            ws.cell(row=current_row, column=column_mapping['wrz']).value = float(wrz)
        
        # 统计
        if polder_code:
            polder_set.add(polder_code)
        if dike_code:
            dike_set.add(dike_code)
        
        current_row += 1
        
        # 显示进度
        if (idx + 1) % 100 == 0 or idx == len(features) - 1:
            print(f"  已处理 {idx + 1}/{len(features)} 条堤段记录...")
    
    total_records = current_row - start_row
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示统计
    print(f"\n✓ 数据统计:")
    print(f"  - 总堤段数: {len(features)} 条")
    print(f"  - 填充记录数: {total_records} 条")
    print(f"  - 涉及保护片: {len(polder_set)} 个")
    print(f"  - 涉及堤防: {len(dike_set)} 个")
    
    return total_records, len(polder_set), len(dike_set)

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
def generate_report(excel_path, geojson_path, filled_count, polder_count, dike_count, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='堤段对应信息', header=1)
    data_df = df.reset_index(drop=True)
    
    # 统计每个保护片的堤段数量
    polder_stats = data_df.groupby('*保护片编码').size().reset_index(name='堤段数量')
    polder_stats = polder_stats.sort_values('堤段数量', ascending=False)
    
    # 统计每个堤防的堤段数量
    dike_stats = data_df.groupby('堤防编码').size().reset_index(name='堤段数量')
    dike_stats = dike_stats.sort_values('堤段数量', ascending=False)
    
    # 统计岸别分布
    zya_stats = data_df['*岸别（L/R）'].value_counts()
    
    report = f"""
{'=' * 80}
堤段对应信息填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

# from xlsx_common import normalize_code, read_geojson, create_backup
一、填充统计
{'-' * 80}
总记录数: {filled_count} 条
涉及保护片: {polder_count} 个
涉及堤防: {dike_count} 个

# from xlsx_common import normalize_code, read_geojson, create_backup
二、岸别分布
{'-' * 80}
{zya_stats.to_string()}

# from xlsx_common import normalize_code, read_geojson, create_backup
三、各保护片的堤段数量（前10个）
{'-' * 80}
{polder_stats.head(10).to_string(index=False)}

# from xlsx_common import normalize_code, read_geojson, create_backup
四、各堤防的堤段数量
{'-' * 80}
{dike_stats.to_string(index=False)}

# from xlsx_common import normalize_code, read_geojson, create_backup
五、字段说明
{'-' * 80}
- river_code: 河道编码（大写）
- river_name: 河道名称
- dike_code: 堤防编码（大写）
- dike_name: 堤防名称
- ds_code: 堤段编码
- ds_name: 堤段名称
- zya: 岸别（L=左岸，R=右岸）
- polder_code: 保护片编码（大写）
- ddgc: 堤顶高程（米）
- lc: 里程（米）
- ds_length: 堤段长度（米）
- lgtd/lttd: 中心点经纬度
- region_code: 区域编码
- drz/grz/wrz: 设计/保证/警戒水位（米）

# from xlsx_common import normalize_code, read_geojson, create_backup
六、数据来源
{'-' * 80}
所有字段均从 GeoJSON properties 中提取
编码字段已自动规范化为大写

# from xlsx_common import normalize_code, read_geojson, create_backup
{'=' * 80}
注意事项:
- 每条堤段记录包含完整的堤防、保护片信息
- 水位信息可能为空（取决于源数据）
- 里程(LC)和堤段长度已包含
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '堤段对应信息填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条堤段记录")
    print(f"✓ 涉及 {polder_count} 个保护片")
    print(f"✓ 涉及 {dike_count} 个堤防")
    print(f"✓ 平均每个保护片有 {filled_count/polder_count:.1f} 条堤段")
    print(f"✓ 平均每个堤防有 {filled_count/dike_count:.1f} 条堤段")
    print("=" * 80)

def main():
    """主函数"""
    # 解析命令行参数
    parser = argparse.ArgumentParser(
        description='堤段对应信息填充 - 从 GeoJSON 导入到风险分析 Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
示例:
  # 使用默认路径（熟溪）
  python3 3.04_risk_dike_section_info.py
  
  # 指定路径（华溪）
  python3 3.04_risk_dike_section_info.py -g /path/to/dd.geojson -e /path/to/hx.xlsx
        '''
    )
    parser.add_argument('-g', '--geojson', help='GeoJSON 文件路径 (dd.geojson)')
    parser.add_argument('-e', '--excel', help='Excel 文件路径 (risk_xx.xlsx)')
    parser.add_argument('-s', '--sheet', help='Sheet 名称', default='堤段对应信息')
    args = parser.parse_args()
    
    # 优先使用命令行参数，否则使用默认值
    geojson_path = args.geojson or 'input/geojson/dd_fix.geojson'
    excel_path = args.excel or 'output/risk_sx.xlsx'
    sheet_name = args.sheet
    
    print("\n" + "=" * 80)
    print("堤段对应信息导入工具 - GeoJSON to Excel (v2.0)")
    print("=" * 80)
    print(f"  GeoJSON: {geojson_path}")
    print(f"  Excel:   {excel_path}")
    print(f"  Sheet:   {sheet_name}")
    print("=" * 80 + "\n")
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        sys.exit(1)
    
    try:
        # 1. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 2. 创建备份
        backup_path = create_backup(excel_path)
        
        # 3. 填充数据
        filled_count, polder_count, dike_count = fill_dike_section_info(geojson_data, excel_path, sheet_name)
        
        # 4. 生成报告
        generate_report(excel_path, geojson_path, filled_count, polder_count, dike_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
if __name__ == '__main__':
    main()

# from xlsx_common import normalize_code, read_geojson, create_backup
