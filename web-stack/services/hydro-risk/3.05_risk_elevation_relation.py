#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
保护片堤段堤顶高程对应关系信息导入脚本 - 从 GeoJSON 导入到风险分析 Excel

【脚本编号】3.05
【目标文件】风险分析-{流域}.xlsx
【目标Sheet】保护片堤段堤顶高程对应关系信息

【使用方法】
    # 使用默认路径
    python3 3.05_*.py
    
    # 指定路径
    python3 3.05_*.py -g /path/to/data.geojson -e /path/to/risk.xlsx

【功能】
    1. 从 input/geojson/dd_fix.geojson 读取堤段数据
    2. 提取保护片、堤段、堤顶高程、里程的对应关系
    3. 参考模板的字段结构
    4. 填充到 output/risk_xx.xlsx 的"保护片堤段堤顶高程对应关系信息" sheet
    5. 自动备份原文件
    6. 生成详细报告

【文件要求】
    - input/geojson/dd_fix.geojson    源数据（熟溪堤段GeoJSON，含river_name）
    - templates/risk_bx.xlsx       白溪模板（参考字段结构）
    - output/risk_sx.xlsx          目标Excel（熟溪）

【依赖安装】
    pip install pandas openpyxl

【字段说明】
    - polder_code: 保护片编码（大写，如SX0003）
    - ds_code: 堤段编码（如20250001）
    - ddgc: 堤顶高程（米）
    - lc: 里程（米）
"""
import argparse
import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    normalize_code,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def fill_elevation_relation(geojson_data, excel_path, sheet_name='保护片堤段堤顶高程对应关系信息'):
    """
    将 GeoJSON 数据填充到 Excel 的保护片堤段堤顶高程对应关系信息 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        sheet_name: sheet 名称
    """
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'polder_code': 1,    # 保护片编码
        'ds_code': 2,        # 堤段编码
        'ddgc': 3,           # 堤顶高程
        'lc': 4              # 里程
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据
    start_row = 3
    features = geojson_data.get('features', [])
    
    print(f"✓ 开始填充数据...")
    
    current_row = start_row
    polder_set = set()  # 用于统计唯一保护片数
    
    for idx, feature in enumerate(features):
        props = feature.get('properties', {})
        
        # 保护片编码 - 规范化为大写
        polder_id = props.get('polder_id', '')
        polder_code = normalize_code(polder_id)
        
        # 堤段编码
        ds_code = props.get('ds_code', '')
        
        # 堤顶高程
        ddgc = props.get('ddgc')
        
        # 里程
        lc = props.get('LC')
        
        # 跳过无效记录（至少需要堤段编码）
        if not ds_code:
            continue
        
        # 填充数据
        ws.cell(row=current_row, column=column_mapping['polder_code']).value = polder_code
        ws.cell(row=current_row, column=column_mapping['ds_code']).value = str(ds_code)
        
        # 数值字段（可能为None）
        if ddgc is not None:
            ws.cell(row=current_row, column=column_mapping['ddgc']).value = float(ddgc)
        
        if lc is not None:
            ws.cell(row=current_row, column=column_mapping['lc']).value = float(lc)
        
        # 统计
        if polder_code:
            polder_set.add(polder_code)
        
        current_row += 1
        
        # 显示进度
        if (idx + 1) % 100 == 0 or idx == len(features) - 1:
            print(f"  已处理 {idx + 1}/{len(features)} 条堤段记录...")
    
    total_records = current_row - start_row
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示统计
    print(f"\n✓ 数据统计:")
    print(f"  - 总堤段数: {len(features)} 条")
    print(f"  - 填充记录数: {total_records} 条")
    print(f"  - 涉及保护片: {len(polder_set)} 个")
    
    return total_records, len(polder_set)

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
def generate_report(excel_path, geojson_path, filled_count, polder_count, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='保护片堤段堤顶高程对应关系信息', header=1)
    data_df = df.reset_index(drop=True)
    
    # 统计每个保护片的堤段数量
    polder_stats = data_df.groupby('*保护片编码').size().reset_index(name='堤段数量')
    polder_stats = polder_stats.sort_values('堤段数量', ascending=False)
    
    # 统计堤顶高程范围
    ddgc_stats = data_df['*堤顶高程'].describe()
    
    # 统计里程范围
    lc_stats = data_df['*里程'].describe()
    
    report = f"""
{'=' * 80}
保护片堤段堤顶高程对应关系信息填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

# from xlsx_common import normalize_code, read_geojson, create_backup
一、填充统计
{'-' * 80}
总记录数: {filled_count} 条
涉及保护片: {polder_count} 个

# from xlsx_common import normalize_code, read_geojson, create_backup
二、各保护片的堤段数量
{'-' * 80}
{polder_stats.to_string(index=False)}

# from xlsx_common import normalize_code, read_geojson, create_backup
三、堤顶高程统计
{'-' * 80}
最小值: {ddgc_stats['min']:.2f} 米
最大值: {ddgc_stats['max']:.2f} 米
平均值: {ddgc_stats['mean']:.2f} 米
中位数: {ddgc_stats['50%']:.2f} 米

# from xlsx_common import normalize_code, read_geojson, create_backup
四、里程统计
{'-' * 80}
最小值: {lc_stats['min']:.0f} 米
最大值: {lc_stats['max']:.0f} 米
平均值: {lc_stats['mean']:.0f} 米
中位数: {lc_stats['50%']:.0f} 米

# from xlsx_common import normalize_code, read_geojson, create_backup
五、字段说明
{'-' * 80}
- polder_code: 保护片编码（大写）
- ds_code: 堤段编码
- ddgc: 堤顶高程（米）
- lc: 里程（米）

# from xlsx_common import normalize_code, read_geojson, create_backup
六、数据来源
{'-' * 80}
- 保护片编码: GeoJSON polder_id字段（规范化为大写）
- 堤段编码: GeoJSON ds_code字段
- 堤顶高程: GeoJSON ddgc字段
- 里程: GeoJSON LC字段

# from xlsx_common import normalize_code, read_geojson, create_backup
{'=' * 80}
注意事项:
- 每条堤段对应一个保护片
- 堤顶高程和里程均为数值型数据
- 编码已自动规范化为大写
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '保护片堤段堤顶高程对应关系填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条对应关系记录")
    print(f"✓ 涉及 {polder_count} 个保护片")
    print(f"✓ 堤顶高程范围: {ddgc_stats['min']:.2f} ~ {ddgc_stats['max']:.2f} 米")
    print(f"✓ 里程范围: {lc_stats['min']:.0f} ~ {lc_stats['max']:.0f} 米")
    print("=" * 80)

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='保护片堤段堤顶高程对应关系填充',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例: python3 3.05_risk_elevation_relation.py -g /path/to/dd.geojson -e /path/to/hx.xlsx'
    )
    parser.add_argument('-g', '--geojson', help='GeoJSON 文件路径 (dd.geojson)')
    parser.add_argument('-e', '--excel', help='Excel 文件路径 (risk_xx.xlsx)')
    parser.add_argument('-s', '--sheet', help='Sheet 名称', default='保护片堤段堤顶高程对应关系信息')
    args = parser.parse_args()
    
    geojson_path = args.geojson or 'input/geojson/dd_fix.geojson'
    excel_path = args.excel or 'output/risk_sx.xlsx'
    sheet_name = args.sheet
    
    print("\n" + "=" * 80)
    print("保护片堤段堤顶高程对应关系信息导入工具 (v2.0)")
    print("=" * 80)
    print(f"  GeoJSON: {geojson_path}")
    print(f"  Excel:   {excel_path}")
    print("=" * 80 + "\n")
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        sys.exit(1)
    
    try:
        # 1. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 2. 创建备份
        backup_path = create_backup(excel_path)
        
        # 3. 填充数据
        filled_count, polder_count = fill_elevation_relation(geojson_data, excel_path, sheet_name)
        
        # 4. 生成报告
        generate_report(excel_path, geojson_path, filled_count, polder_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
if __name__ == '__main__':
    main()

# from xlsx_common import normalize_code, read_geojson, create_backup
