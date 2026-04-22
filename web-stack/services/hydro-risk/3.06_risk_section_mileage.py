#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
断面里程对应关系信息导入脚本 - 从 GeoJSON 导入到风险分析 Excel

【脚本编号】3.06
【目标文件】风险分析-{流域}.xlsx
【目标Sheet】断面里程对应关系信息

【使用方法】
    # 使用默认路径
    python3 3.06_*.py
    
    # 指定路径
    python3 3.06_*.py -g /path/to/data.geojson -e /path/to/risk.xlsx

【功能】
    1. 从 input/dm_LC.geojson 读取断面数据
    2. 提取断面编码和里程的对应关系
    3. 参考模板的字段结构
    4. 按里程排序，生成顺序编号作为断面编码
    5. 填充到 output/risk_xx.xlsx 的"断面里程对应关系信息" sheet
    6. 自动备份原文件
    7. 生成详细报告

【文件要求】
    - input/dm_LC.geojson    源数据（熟溪断面GeoJSON）
    - templates/risk_bx.xlsx    白溪模板（参考字段结构）
    - output/risk_sx.xlsx       目标Excel（熟溪）

【依赖安装】
    pip install pandas openpyxl

【字段说明】
    - dot_code: 断面编码（顺序编号，从1开始）
    - lc: 里程（米）
"""
import argparse
import pandas as pd
import json
import re
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    natural_sort_key,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def fill_section_mileage(geojson_data, excel_path, sheet_name='断面里程对应关系信息'):
    """
    将 GeoJSON 数据填充到 Excel 的断面里程对应关系信息 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        sheet_name: sheet 名称
    """
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'dot_code': 1,    # 断面编码
        'lc': 2           # 里程
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 提取断面数据
    features = geojson_data.get('features', [])
    sections = []
    
    for feature in features:
        props = feature.get('properties', {})
        lc = props.get('LC')
        dot_name = props.get('断面编')  # 提取断面编号（如 sx1, sx2）
        
        if lc is not None:
            sections.append({
                'lc': float(lc),
                'dot_name': dot_name
            })
    
    # 按断面编码排序（自然排序，支持sx1, sx2, sx10, gzx1等）
    sections_sorted = sorted(sections, key=lambda x: natural_sort_key(x.get('dot_name', '')))
    
    print(f"✓ 开始填充数据...")
    print(f"  共 {len(sections_sorted)} 条断面记录，按断面编码排序")
    
    # 从第3行开始填充数据
    start_row = 3
    current_row = start_row
    
    for idx, section in enumerate(sections_sorted):
        # 使用 GeoJSON 中的断面编号（如 sx1, sx2），如果没有则使用序号
        dot_code = section.get('dot_name', idx + 1)
        lc = section['lc']
        
        # 填充数据
        ws.cell(row=current_row, column=column_mapping['dot_code']).value = dot_code
        ws.cell(row=current_row, column=column_mapping['lc']).value = lc
        
        current_row += 1
        
        # 显示进度
        if (idx + 1) % 10 == 0 or idx == len(sections_sorted) - 1:
            print(f"  已处理 {idx + 1}/{len(sections_sorted)} 条断面记录...")
    
    total_records = current_row - start_row
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示统计
    print(f"\n✓ 数据统计:")
    print(f"  - 填充记录数: {total_records} 条")
    print(f"  - 断面编码: 使用实际断面名称（sx1, sx2...）")
    
    return total_records

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
def generate_report(excel_path, geojson_path, filled_count, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='断面里程对应关系信息', header=1)
    data_df = df.reset_index(drop=True)
    
    # 统计里程范围
    lc_stats = data_df['*里程'].describe()
    
    report = f"""
{'=' * 80}
断面里程对应关系信息填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

# from xlsx_common import normalize_code, read_geojson, create_backup
一、填充统计
{'-' * 80}
总记录数: {filled_count} 条
断面编码范围: 1 ~ {filled_count}

# from xlsx_common import normalize_code, read_geojson, create_backup
二、里程统计
{'-' * 80}
最小值: {lc_stats['min']:.0f} 米
最大值: {lc_stats['max']:.0f} 米
平均值: {lc_stats['mean']:.0f} 米
中位数: {lc_stats['50%']:.0f} 米

# from xlsx_common import normalize_code, read_geojson, create_backup
三、字段说明
{'-' * 80}
- dot_code: 断面编码（使用 GeoJSON 中的"断面编"字段，如 sx1, sx2）
- lc: 里程（米）

# from xlsx_common import normalize_code, read_geojson, create_backup
四、数据来源
{'-' * 80}
- 断面编码: GeoJSON "断面编"字段（如 sx1, sx2, sx3...）
- 里程: GeoJSON LC字段

# from xlsx_common import normalize_code, read_geojson, create_backup
{'=' * 80}
注意事项:
- 断面按里程从小到大排序
- 断面编码使用实际断面名称（sx1, sx2, sx3...）
- 与"主要断面基础信息"中的断面名称一致
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '断面里程对应关系填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条断面里程对应关系")
    print(f"✓ 断面编码: 使用实际断面名称（sx1 ~ sx{filled_count}）")
    print(f"✓ 里程范围: {lc_stats['min']:.0f} ~ {lc_stats['max']:.0f} 米")
    print("=" * 80)

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='断面里程对应关系填充',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例: python3 3.06_risk_section_mileage.py -g /path/to/dm_LC.geojson -e /path/to/hx.xlsx'
    )
    parser.add_argument('-g', '--geojson', help='GeoJSON 文件路径 (dm_LC.geojson)')
    parser.add_argument('-e', '--excel', help='Excel 文件路径 (risk_xx.xlsx)')
    parser.add_argument('-s', '--sheet', help='Sheet 名称', default='断面里程对应关系信息')
    args = parser.parse_args()
    
    geojson_path = args.geojson or 'input/dm_LC.geojson'
    excel_path = args.excel or 'output/risk_sx.xlsx'
    sheet_name = args.sheet
    
    print("\n" + "=" * 80)
    print("断面里程对应关系信息导入工具 (v2.0)")
    print("=" * 80)
    print(f"  GeoJSON: {geojson_path}")
    print(f"  Excel:   {excel_path}")
    print("=" * 80 + "\n")
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        sys.exit(1)
    
    try:
        # 1. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 2. 创建备份
        backup_path = create_backup(excel_path)
        
        # 3. 填充数据
        filled_count = fill_section_mileage(geojson_data, excel_path, sheet_name)
        
        # 4. 生成报告
        generate_report(excel_path, geojson_path, filled_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
if __name__ == '__main__':
    main()

# from xlsx_common import normalize_code, read_geojson, create_backup
