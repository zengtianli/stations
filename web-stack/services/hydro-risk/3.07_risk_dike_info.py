#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
堤防信息导入脚本 - 从 GeoJSON 导入到风险分析 Excel

【脚本编号】3.07
【目标文件】风险分析-{流域}.xlsx
【目标Sheet】堤防信息

【使用方法】
    # 使用默认路径
    python3 3.07_*.py
    
    # 指定路径
    python3 3.07_*.py -g /path/to/data.geojson -e /path/to/risk.xlsx

【功能】
    1. 从 input/df_with_elevation_lc.geojson 读取堤防数据
    2. 提取堤防详细信息（编码、名称、起终点、长度、坐标等）
    3. 参考模板的字段结构
    4. 填充到 output/risk_xx.xlsx 的"堤防信息" sheet
    5. 自动备份原文件
    6. 生成详细报告

【文件要求】
    - input/df_with_elevation_lc.geojson  源数据（熟溪堤防GeoJSON，含高程里程）
    - templates/risk_bx.xlsx               白溪模板（参考字段结构）
    - output/risk_sx.xlsx                  目标Excel（熟溪）

【依赖安装】
    pip install pandas openpyxl

【字段说明】
    - river_code: 河流编码（大写，如SX）
    - river_name: 河流名称（如熟溪）
    - dike_code: 堤防编码（大写，如SXDF0001）
    - dike_name: 堤防名称（如熟溪堤防0001）
    - start_name: 起点站
    - end_name: 终点站
    - zya: 岸别（L/R）
    - polder_code: 保护片编码（大写）
    - qdgc: 起点高程（米）
    - zdgc: 终点高程（米）
    - qdlc: 起点里程（米）
    - zdlc: 终点里程（米）
    - ds_length: 堤段长度（米）
    - lgtd: 经度
    - lttd: 维度
    - geom: 空间数据（空）
    - design_type: 设计标准
    - addvnm: 所属区域
    - addvcd: 区域编码
"""
import argparse
import pandas as pd
import json
import re
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    normalize_code,
    extract_dike_number,
    generate_dike_code,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def fill_dike_info(geojson_data, excel_path, sheet_name='堤防信息'):
    """
    将 GeoJSON 数据填充到 Excel 的堤防信息 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        sheet_name: sheet 名称
    """
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'river_code': 1,     # 河流编码
        'river_name': 2,     # 河流名称
        'dike_code': 3,      # 堤防编码
        'dike_name': 4,      # 堤防名称
        'start_name': 5,     # 起点站
        'end_name': 6,       # 终点站
        'zya': 7,            # 岸别
        'polder_code': 8,    # 保护片编码
        'qdgc': 9,           # 起点高程
        'zdgc': 10,          # 终点高程
        'qdlc': 11,          # 起点里程
        'zdlc': 12,          # 终点里程
        'ds_length': 13,     # 堤段长度
        'lgtd': 14,          # 经度
        'lttd': 15,          # 维度
        'geom': 16,          # 空间数据
        'design_type': 17,   # 设计标准
        'addvnm': 18,        # 所属区域
        'addvcd': 19         # 区域编码
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据
    start_row = 3
    features = geojson_data.get('features', [])
    
    print(f"✓ 开始填充数据...")
    
    current_row = start_row
    
    for idx, feature in enumerate(features):
        props = feature.get('properties', {})
        
        # 河流编码和名称 - 规范化为大写
        river_code = normalize_code(props.get('subCode', ''))  # 支流编码作为河流编码
        river_name = props.get('riverName', '')
        
        # 堤防编码和名称 - 直接使用源数据的编码（如 SXDF0003）
        dike_name = props.get('dikeName', '')
        dike_code = normalize_code(props.get('dikeCode', ''))
        
        # 起终点
        start_name = props.get('startName', '')
        end_name = props.get('endName', '')
        
        # 岸别
        zya = props.get('zya', '').upper() if props.get('zya') else ''
        
        # 保护片编码 - 规范化为大写
        polder_code = normalize_code(props.get('polderCode', ''))
        
        # 高程
        qdgc = props.get('qdgc')  # 起点高程
        zdgc = props.get('zdgc')  # 终点高程
        
        # 里程
        qdlc = props.get('qdlc')  # 起点里程
        zdlc = props.get('zdlc')  # 终点里程
        
        # 长度
        ds_length = props.get('dsLength')
        
        # 经纬度
        lgtd = props.get('lgtd')
        lttd = props.get('lttd')
        
        # 设计标准
        design_type = props.get('type')
        
        # 所属区域
        addvnm = props.get('adcdName', '')
        addvcd = props.get('adcdCode', '')
        
        # 填充数据
        ws.cell(row=current_row, column=column_mapping['river_code']).value = river_code
        ws.cell(row=current_row, column=column_mapping['river_name']).value = river_name
        ws.cell(row=current_row, column=column_mapping['dike_code']).value = dike_code
        ws.cell(row=current_row, column=column_mapping['dike_name']).value = dike_name
        ws.cell(row=current_row, column=column_mapping['start_name']).value = start_name
        ws.cell(row=current_row, column=column_mapping['end_name']).value = end_name
        ws.cell(row=current_row, column=column_mapping['zya']).value = zya
        ws.cell(row=current_row, column=column_mapping['polder_code']).value = polder_code
        
        # 数值字段
        if qdgc is not None:
            ws.cell(row=current_row, column=column_mapping['qdgc']).value = float(qdgc)
        
        if zdgc is not None:
            ws.cell(row=current_row, column=column_mapping['zdgc']).value = float(zdgc)
        
        if qdlc is not None:
            ws.cell(row=current_row, column=column_mapping['qdlc']).value = float(qdlc)
        
        if zdlc is not None:
            ws.cell(row=current_row, column=column_mapping['zdlc']).value = float(zdlc)
        
        if ds_length is not None:
            ws.cell(row=current_row, column=column_mapping['ds_length']).value = float(ds_length)
        
        if lgtd is not None:
            ws.cell(row=current_row, column=column_mapping['lgtd']).value = float(lgtd)
        
        if lttd is not None:
            ws.cell(row=current_row, column=column_mapping['lttd']).value = float(lttd)
        
        # geom 留空
        ws.cell(row=current_row, column=column_mapping['geom']).value = ''
        
        # 设计标准
        if design_type is not None:
            ws.cell(row=current_row, column=column_mapping['design_type']).value = design_type
        
        # 所属区域
        ws.cell(row=current_row, column=column_mapping['addvnm']).value = addvnm
        ws.cell(row=current_row, column=column_mapping['addvcd']).value = addvcd
        
        current_row += 1
        
        # 显示进度
        if (idx + 1) % 5 == 0 or idx == len(features) - 1:
            print(f"  已处理 {idx + 1}/{len(features)} 条堤防记录...")
    
    total_records = current_row - start_row
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示统计
    print(f"\n✓ 数据统计:")
    print(f"  - 填充记录数: {total_records} 条")
    
    return total_records

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
def generate_report(excel_path, geojson_path, filled_count, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='堤防信息', header=1)
    data_df = df.reset_index(drop=True)
    
    report = f"""
{'=' * 80}
堤防信息填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

# from xlsx_common import normalize_code, read_geojson, create_backup
一、填充统计
{'-' * 80}
总记录数: {filled_count} 条

# from xlsx_common import normalize_code, read_geojson, create_backup
二、字段说明
{'-' * 80}
- river_code/river_name: 河流编码/名称
- dike_code/dike_name: 堤防编码/名称（已规范化为大写）
- start_name/end_name: 起点站/终点站
- zya: 岸别（L/R）
- polder_code: 保护片编码（已规范化为大写）
- qdgc/zdgc: 起点/终点高程（米）
- qdlc/zdlc: 起点/终点里程（米）
- ds_length: 堤段长度（米）
- lgtd/lttd: 经度/维度
- design_type: 设计标准
- addvnm/addvcd: 所属区域/区域编码

# from xlsx_common import normalize_code, read_geojson, create_backup
三、数据来源
{'-' * 80}
所有字段均从 GeoJSON properties 中提取
编码字段已自动规范化为大写

# from xlsx_common import normalize_code, read_geojson, create_backup
{'=' * 80}
注意事项:
- geom 字段保持为空
- 编码已自动规范化为大写
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '堤防信息填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条堤防信息")
    print("=" * 80)

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='堤防信息填充',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例: python3 3.07_risk_dike_info.py -g /path/to/df.geojson -e /path/to/hx.xlsx'
    )
    parser.add_argument('-g', '--geojson', help='GeoJSON 文件路径 (df.geojson)')
    parser.add_argument('-e', '--excel', help='Excel 文件路径 (risk_xx.xlsx)')
    parser.add_argument('-s', '--sheet', help='Sheet 名称', default='堤防信息')
    args = parser.parse_args()
    
    geojson_path = args.geojson or 'input/df_with_elevation_lc.geojson'
    excel_path = args.excel or 'output/risk_sx.xlsx'
    sheet_name = args.sheet
    
    print("\n" + "=" * 80)
    print("堤防信息导入工具 (v2.0)")
    print("=" * 80)
    print(f"  GeoJSON: {geojson_path}")
    print(f"  Excel:   {excel_path}")
    print("=" * 80 + "\n")
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        sys.exit(1)
    
    try:
        # 1. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 2. 创建备份
        backup_path = create_backup(excel_path)
        
        # 3. 填充数据
        filled_count = fill_dike_info(geojson_data, excel_path, sheet_name)
        
        # 4. 生成报告
        generate_report(excel_path, geojson_path, filled_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
if __name__ == '__main__':
    main()

# from xlsx_common import normalize_code, read_geojson, create_backup
