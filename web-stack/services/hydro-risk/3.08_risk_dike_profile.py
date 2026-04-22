#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
堤防纵剖面数据导入脚本 - 从 GeoJSON 导入到风险分析 Excel

【脚本编号】3.08
【目标文件】风险分析-{流域}.xlsx
【目标Sheet】堤防纵剖面数据

【使用方法】
    # 使用默认路径
    python3 3.08_*.py
    
    # 指定路径
    python3 3.08_*.py -g /path/to/data.geojson -e /path/to/risk.xlsx

【功能】
    1. 从 input/geojson/dd_fix.geojson 读取堤段数据
    2. 提取河道编码、里程、河底高程、左右岸高程等信息
    3. 参考模板的字段结构
    4. 填充到 output/risk_xx.xlsx 的"堤防纵剖面数据" sheet
    5. 自动备份原文件
    6. 生成详细报告

【文件要求】
    - input/geojson/dd_fix.geojson    源数据（熟溪堤段GeoJSON）
    - templates/risk_bx.xlsx           白溪模板（参考字段结构）
    - output/risk_sx.xlsx              目标Excel（熟溪）

【依赖安装】
    pip install pandas openpyxl

【字段说明】
    - river_code: 河道编码（大写，如SX）
    - lc: 里程（米）
    - river_bottom_elevation: 河底高程（米）
    - river_left_elevation: 左岸高程（米）
    - river_right_elevation: 右岸高程（米）
    - river_plotline: 是否河道主线（空）
"""
import argparse
import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path

# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    normalize_code,
)
from file_ops import (
    read_geojson,
    create_backup,
)


def fill_profile_data(geojson_data, excel_path, sheet_name='堤防纵剖面数据'):
    """
    将 GeoJSON 数据填充到 Excel 的堤防纵剖面数据 sheet
    
    Args:
        geojson_data: GeoJSON 数据
        excel_path: Excel 文件路径
        sheet_name: sheet 名称
    """
    wb = load_workbook(excel_path)
    
    if sheet_name not in wb.sheetnames:
        print(f"✗ 错误: 未找到 sheet '{sheet_name}'")
        print(f"  可用的 sheets: {', '.join(wb.sheetnames)}")
        sys.exit(1)
    
    ws = wb[sheet_name]
    
    # 列名映射（Excel 列索引从 1 开始）
    column_mapping = {
        'river_code': 1,              # 河道编码
        'lc': 2,                      # 里程
        'river_bottom_elevation': 3,  # 河底高程
        'river_left_elevation': 4,    # 左岸高程
        'river_right_elevation': 5,   # 右岸高程
        'river_plotline': 6           # 是否河道主线
    }
    
    # 清除从第3行开始的所有数据（保留表头和说明行）
    print(f"✓ 清除 sheet '{sheet_name}' 中的旧数据...")
    max_row = ws.max_row
    for row in range(3, max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None
    
    # 从第3行开始填充数据
    start_row = 3
    features = geojson_data.get('features', [])
    
    print(f"✓ 开始填充数据...")
    
    # 按(河道编码, 里程)分组，分别记录左右岸高程
    lc_data = {}
    
    for feature in features:
        props = feature.get('properties', {})
        
        # 河道编码
        river_code = normalize_code(props.get('river_code', ''))
        
        # 里程
        lc = props.get('LC')
        
        # 左右岸标识
        zya = props.get('zya', '').strip().upper()
        
        if lc is None or not river_code:
            continue
        
        # 使用(河道编码, 里程)作为组合键，避免不同河道相同里程数据覆盖
        lc_key = (river_code, int(lc))
        
        if lc_key not in lc_data:
            lc_data[lc_key] = {
                'river_code': river_code,
                'lc': lc,
                'river_bottom_elevation': None,  # 河底高程
                'left_elevation': None,          # 左岸高程
                'right_elevation': None          # 右岸高程
            }
        
        # 根据zya字段，分别记录左右岸高程
        ddgc = props.get('ddgc')
        if ddgc is not None:
            # 河底高程取平均值或第一个值
            if lc_data[lc_key]['river_bottom_elevation'] is None:
                lc_data[lc_key]['river_bottom_elevation'] = float(ddgc)
            
            # 根据左右岸标识，分别记录
            if zya == 'L':  # 左岸
                if lc_data[lc_key]['left_elevation'] is None:
                    lc_data[lc_key]['left_elevation'] = float(ddgc)
            elif zya == 'R':  # 右岸
                if lc_data[lc_key]['right_elevation'] is None:
                    lc_data[lc_key]['right_elevation'] = float(ddgc)
    
    # 按(河道编码, 里程)排序
    sorted_lcs = sorted(lc_data.keys(), key=lambda x: (x[0], x[1]))
    
    # 对左右岸高程进行插值处理
    def interpolate_elevations(lc_data, sorted_lcs, field_name):
        """
        对高程数据进行插值
        - 头尾空值：使用最近的非空值填充
        - 中间空值：使用线性插值
        """
        elevations = [lc_data[lc].get(field_name) for lc in sorted_lcs]
        
        # 处理头部空值
        first_non_null_idx = None
        for i, val in enumerate(elevations):
            if val is not None:
                first_non_null_idx = i
                break
        
        if first_non_null_idx is not None and first_non_null_idx > 0:
            # 用第一个非空值填充头部
            for i in range(first_non_null_idx):
                elevations[i] = elevations[first_non_null_idx]
                lc_data[sorted_lcs[i]][field_name] = elevations[first_non_null_idx]
        
        # 处理尾部空值
        last_non_null_idx = None
        for i in range(len(elevations) - 1, -1, -1):
            if elevations[i] is not None:
                last_non_null_idx = i
                break
        
        if last_non_null_idx is not None and last_non_null_idx < len(elevations) - 1:
            # 用最后一个非空值填充尾部
            for i in range(last_non_null_idx + 1, len(elevations)):
                elevations[i] = elevations[last_non_null_idx]
                lc_data[sorted_lcs[i]][field_name] = elevations[last_non_null_idx]
        
        # 处理中间空值（线性插值）
        i = 0
        while i < len(elevations):
            if elevations[i] is None:
                # 找到空值段的起始和结束
                start_idx = i - 1
                end_idx = i
                
                # 找到下一个非空值
                while end_idx < len(elevations) and elevations[end_idx] is None:
                    end_idx += 1
                
                # 如果找到了前后非空值，进行线性插值
                if start_idx >= 0 and end_idx < len(elevations):
                    start_val = elevations[start_idx]
                    end_val = elevations[end_idx]
                    # 获取里程值：如果键是元组(river_code, lc)，取第二个元素；否则直接使用
                    start_lc = sorted_lcs[start_idx][1] if isinstance(sorted_lcs[start_idx], tuple) else sorted_lcs[start_idx]
                    end_lc = sorted_lcs[end_idx][1] if isinstance(sorted_lcs[end_idx], tuple) else sorted_lcs[end_idx]
                    
                    # 线性插值
                    for j in range(start_idx + 1, end_idx):
                        curr_lc = sorted_lcs[j][1] if isinstance(sorted_lcs[j], tuple) else sorted_lcs[j]
                        ratio = (curr_lc - start_lc) / (end_lc - start_lc)
                        interpolated_val = start_val + ratio * (end_val - start_val)
                        elevations[j] = interpolated_val
                        lc_data[sorted_lcs[j]][field_name] = interpolated_val
                
                i = end_idx
            else:
                i += 1
    
    print(f"✓ 开始插值处理...")
    # 按河道分组进行插值
    river_groups = {}
    for lc_key in sorted_lcs:
        river_code = lc_key[0]
        if river_code not in river_groups:
            river_groups[river_code] = []
        river_groups[river_code].append(lc_key)
    
    # 对每个河道分别进行插值
    for river_code, river_lcs in river_groups.items():
        print(f"  处理河道 {river_code} 的插值...")
        interpolate_elevations(lc_data, river_lcs, 'left_elevation')
        interpolate_elevations(lc_data, river_lcs, 'right_elevation')
    
    print(f"✓ 插值处理完成")
    
    current_row = start_row
    
    for idx, lc_key in enumerate(sorted_lcs):
        data = lc_data[lc_key]
        
        river_code = data['river_code']
        lc = data['lc']
        river_bottom_elevation = data['river_bottom_elevation']
        river_left_elevation = data['left_elevation']
        river_right_elevation = data['right_elevation']
        
        # 河底高程：假设比堤顶低一定高度（取左右岸平均值）
        if river_bottom_elevation is not None:
            river_bottom_elevation = river_bottom_elevation - 5.0
        elif river_left_elevation is not None and river_right_elevation is not None:
            avg_elevation = (river_left_elevation + river_right_elevation) / 2.0
            river_bottom_elevation = avg_elevation - 5.0
        elif river_left_elevation is not None:
            river_bottom_elevation = river_left_elevation - 5.0
        elif river_right_elevation is not None:
            river_bottom_elevation = river_right_elevation - 5.0
        
        # 填充数据
        ws.cell(row=current_row, column=column_mapping['river_code']).value = river_code
        
        if lc is not None:
            ws.cell(row=current_row, column=column_mapping['lc']).value = float(lc)
        
        if river_bottom_elevation is not None:
            ws.cell(row=current_row, column=column_mapping['river_bottom_elevation']).value = river_bottom_elevation
        
        if river_left_elevation is not None:
            ws.cell(row=current_row, column=column_mapping['river_left_elevation']).value = river_left_elevation
        
        if river_right_elevation is not None:
            ws.cell(row=current_row, column=column_mapping['river_right_elevation']).value = river_right_elevation
        
        # river_plotline 留空
        ws.cell(row=current_row, column=column_mapping['river_plotline']).value = ''
        
        current_row += 1
        
        # 显示进度
        if (idx + 1) % 100 == 0 or idx == len(sorted_lcs) - 1:
            print(f"  已处理 {idx + 1}/{len(sorted_lcs)} 条里程记录...")
    
    total_records = current_row - start_row
    
    # 保存文件
    wb.save(excel_path)
    print(f"✓ 成功保存文件: {excel_path}")
    
    # 显示统计
    print(f"\n✓ 数据统计:")
    print(f"  - 填充记录数: {total_records} 条")
    print(f"  - 里程范围: {sorted_lcs[0]} ~ {sorted_lcs[-1]} 米")
    
    return total_records

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
def generate_report(excel_path, geojson_path, filled_count, backup_path):
    """生成填充报告"""
    df = pd.read_excel(excel_path, sheet_name='堤防纵剖面数据', header=1)
    data_df = df.reset_index(drop=True)
    
    # 统计高程范围
    bottom_stats = data_df['河底高程'].describe()
    left_stats = data_df['左岸高程'].describe()
    right_stats = data_df['右岸高程'].describe()
    
    report = f"""
{'=' * 80}
堤防纵剖面数据填充报告
{'=' * 80}
填充时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
数据来源: {geojson_path}
目标文件: {excel_path}
备份文件: {backup_path}

# from xlsx_common import normalize_code, read_geojson, create_backup
一、填充统计
{'-' * 80}
总记录数: {filled_count} 条

# from xlsx_common import normalize_code, read_geojson, create_backup
二、高程统计
{'-' * 80}
河底高程:
  最小值: {bottom_stats['min']:.2f} 米
  最大值: {bottom_stats['max']:.2f} 米
  平均值: {bottom_stats['mean']:.2f} 米

# from xlsx_common import normalize_code, read_geojson, create_backup
左岸高程:
  最小值: {left_stats['min']:.2f} 米
  最大值: {left_stats['max']:.2f} 米
  平均值: {left_stats['mean']:.2f} 米

# from xlsx_common import normalize_code, read_geojson, create_backup
右岸高程:
  最小值: {right_stats['min']:.2f} 米
  最大值: {right_stats['max']:.2f} 米
  平均值: {right_stats['mean']:.2f} 米

# from xlsx_common import normalize_code, read_geojson, create_backup
三、字段说明
{'-' * 80}
- river_code: 河道编码（大写）
- lc: 里程（米）
- river_bottom_elevation: 河底高程（米，计算值 = 堤顶高程 - 5）
- river_left_elevation: 左岸高程（米，使用堤顶高程）
- river_right_elevation: 右岸高程（米，使用堤顶高程）
- river_plotline: 是否河道主线（空）

# from xlsx_common import normalize_code, read_geojson, create_backup
四、数据来源
{'-' * 80}
- 河道编码、里程: GeoJSON 字段
- 河底高程: 根据堤顶高程计算（堤顶 - 5米）
- 左右岸高程: 使用堤顶高程

# from xlsx_common import normalize_code, read_geojson, create_backup
{'=' * 80}
注意事项:
- 河底高程为估算值（堤顶高程 - 5米）
- 左右岸高程使用堤顶高程
- 按里程排序并去重
{'=' * 80}
"""
    
    # 保存报告
    report_dir = os.path.dirname(excel_path)
    report_path = os.path.join(report_dir, '堤防纵剖面数据填充报告.txt')
    
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n✓ 报告已保存至: {report_path}")
    
    # 打印汇总信息
    print("\n" + "=" * 80)
    print("填充完成汇总")
    print("=" * 80)
    print(f"✓ 共填充 {filled_count} 条纵剖面数据")
    print(f"✓ 河底高程范围: {bottom_stats['min']:.2f} ~ {bottom_stats['max']:.2f} 米")
    print(f"✓ 左岸高程范围: {left_stats['min']:.2f} ~ {left_stats['max']:.2f} 米")
    print(f"✓ 右岸高程范围: {right_stats['min']:.2f} ~ {right_stats['max']:.2f} 米")
    print("=" * 80)

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='堤防纵剖面数据填充',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例: python3 3.08_risk_dike_profile.py -g /path/to/dd.geojson -e /path/to/hx.xlsx'
    )
    parser.add_argument('-g', '--geojson', help='GeoJSON 文件路径 (dd.geojson)')
    parser.add_argument('-e', '--excel', help='Excel 文件路径 (risk_xx.xlsx)')
    parser.add_argument('-s', '--sheet', help='Sheet 名称', default='堤防纵剖面数据')
    args = parser.parse_args()
    
    geojson_path = args.geojson or 'input/geojson/dd_fix.geojson'
    excel_path = args.excel or 'output/risk_sx.xlsx'
    sheet_name = args.sheet
    
    print("\n" + "=" * 80)
    print("堤防纵剖面数据导入工具 (v2.0)")
    print("=" * 80)
    print(f"  GeoJSON: {geojson_path}")
    print(f"  Excel:   {excel_path}")
    print("=" * 80 + "\n")
    
    # 检查文件是否存在
    if not os.path.exists(geojson_path):
        print(f"✗ 错误: 找不到 GeoJSON 文件: {geojson_path}")
        sys.exit(1)
    
    if not os.path.exists(excel_path):
        print(f"✗ 错误: 找不到 Excel 文件: {excel_path}")
        sys.exit(1)
    
    try:
        # 1. 读取 GeoJSON 数据
        geojson_data = read_geojson(geojson_path)
        
        # 2. 创建备份
        backup_path = create_backup(excel_path)
        
        # 3. 填充数据
        filled_count = fill_profile_data(geojson_data, excel_path, sheet_name)
        
        # 4. 生成报告
        generate_report(excel_path, geojson_path, filled_count, backup_path)
        
        print("\n✓ 所有操作完成！\n")
        
    except Exception as e:
        print(f"\n✗ 发生错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

# from xlsx_common import normalize_code, read_geojson, create_backup

# from xlsx_common import normalize_code, read_geojson, create_backup
if __name__ == '__main__':
    main()

# from xlsx_common import normalize_code, read_geojson, create_backup
