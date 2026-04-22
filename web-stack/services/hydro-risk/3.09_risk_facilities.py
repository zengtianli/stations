#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
==============================================
填充风险分析表中的所有设施相关 sheet
==============================================

【脚本编号】3.09
【目标文件】风险分析-{流域}.xlsx
【目标Sheet】医院、学校、公园景点、敬老院、政府机构、危化企业 重要设施基础信息、
            移动基站、兴趣点、电站、水厂、重要设施（共11个Sheet）

【使用方法】
    # 使用默认路径
    python3 3.09_*.py
    
    # 指定路径
    python3 3.09_*.py -g /path/to/data.geojson -e /path/to/risk.xlsx

【功能说明】
    从 baohu.geojson 读取重要设施数据，填充到 risk_xx.xlsx 的多个 sheet

【输入文件】
    - input/baohu/baohu.geojson：重要设施数据
    - input/region_name_code.csv：区域名称和编码对照表
    
【输出文件】
    - output/risk_xx.xlsx：风险分析表（更新）
    - output/设施填充报告.txt：填充情况报告

【数据映射】
    设施类型（type 字段）：
        0: 避灾场所
        1: 学校
        2: 医院
        3: 水厂
        4: 危化企业
        5: 敬老院
        6: 兴趣点
        7: 公园景点
        8: 电站
        9: 移动基站
        10: 政府部门
"""

import argparse
import pandas as pd
import json
from openpyxl import load_workbook
from datetime import datetime
import os
import sys
from pathlib import Path



# 导入公共工具模块
SCRIPT_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(SCRIPT_DIR / "lib"))

from hydraulic import (
    normalize_code,
)
from file_ops import (
    create_backup,
)

def load_geojson(geojson_path):
    """加载 GeoJSON 文件"""
    print(f"📖 正在读取 GeoJSON: {geojson_path}")
    
    if not os.path.exists(geojson_path):
        print(f"❌ 错误: 文件不存在 - {geojson_path}")
        return None
    
    try:
        with open(geojson_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        feature_count = len(data.get('features', []))
        print(f"✅ 成功读取 {feature_count} 条设施数据")
        return data
    
    except Exception as e:
        print(f"❌ 读取 GeoJSON 失败: {e}")
        return None

def load_region_codes(csv_path):
    """加载区域编码对照表"""
    print(f"📖 正在读取区域编码: {csv_path}")
    
    if not os.path.exists(csv_path):
        print(f"⚠️  警告: 文件不存在 - {csv_path}")
        return {}
    
    try:
        df = pd.read_csv(csv_path)
        # 创建名称到编码的映射
        name_to_code = dict(zip(df['region_name'], df['region_code']))
        print(f"✅ 成功加载 {len(name_to_code)} 条区域编码")
        return name_to_code
    
    except Exception as e:
        print(f"❌ 读取区域编码失败: {e}")
        return {}

    """创建文件备份"""
    if os.path.exists(file_path):
        backup_path = file_path.replace('.xlsx', f'_备份_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        import shutil
        shutil.copy2(file_path, backup_path)
        print(f"💾 已创建备份: {backup_path}")

def get_county_code(county_name, region_codes):
    """获取县级区划编码"""
    return region_codes.get(county_name, '')

def get_town_code(town_name, region_codes):
    """获取乡镇编码"""
    return region_codes.get(town_name, '')

# ============ 主要填充函数 ============

def fill_facilities_data(geojson_data, region_codes, excel_path):
    """
    填充所有设施相关的 sheet
    
    参数：
        geojson_data: GeoJSON 数据
        region_codes: 区域编码映射
        excel_path: Excel 文件路径
    """
    
    # 设施类型映射
    FACILITY_TYPE_MAP = {
        0: '避灾场所',
        1: '学校',
        2: '医院',
        3: '水厂',
        4: '危化企业',
        5: '敬老院',
        6: '兴趣点',
        7: '公园景点',
        8: '电站',
        9: '移动基站',
        10: '政府部门'
    }
    
    # Sheet 配置 - 每个 sheet 的列映射
    SHEET_CONFIGS = {
        '危化企业 重要设施基础信息': {
            'filter_type': [4],
            'columns': {
                '区县级名称': 'addvnm',
                '区县级区划编码': 'addvcd',
                '乡镇编码': 'region_code',
                '乡镇名称': 'region_name',
                '重要设施编码': 'ssid',
                '重要设施名称': 'ssnm',
                '详细地址': 'address',
                '保护片编码': 'polder_code',
                '设施高程': 'elevation',
                '法人': 'legal_person',
                '法人联系方式': 'legal_tel',
                '安全责任人': 'safety_man',
                '安全责任人联系方式': 'safety_tel',
                '经度': 'lng',
                '纬度': 'lat'
            }
        },
        '移动基站': {
            'filter_type': [9],
            'columns': {
                '区县级名称': 'addvnm',
                '区县级区划编码': 'addvcd',
                '乡镇编码': 'region_code',
                '乡镇名称': 'region_name',
                '重要设施编码': 'ssid',
                '重要设施名称': 'ssnm',
                '保护片编码': 'polder_code',
                '设施高程': 'elevation',
                '详细地址': 'address',
                '联系人': 'link_man',
                '联系电话': 'link_tel',
                '基站类型': 'station_type',
                '经度': 'lng',
                '纬度': 'lat'
            }
        },
        '医院': {
            'filter_type': [2],
            'columns': {
                '区县级名称': 'addvnm',
                '区县级区划编码': 'addvcd',
                '乡镇编码': 'region_code',
                '乡镇名称': 'region_name',
                '设施id': 'ssid',
                '重要设施名称': 'ssnm',
                '保护片编码': 'polder_code',
                '设施高程': 'elevation',
                '直接联系人': 'link_man',
                '经度': 'lng',
                '纬度': 'lat',
                '联系人电话': 'link_tel',
                '详细地址': 'address'
            }
        },
        '兴趣点': {
            'filter_type': [6],
            'columns': {
                '区县级名称': 'addvnm',
                '区县级区划编码': 'addvcd',
                '乡镇编码': 'region_code',
                '乡镇名称': 'region_name',
                '设施id': 'ssid',
                '重要设施名称': 'ssnm',
                '保护片编码': 'polder_code',
                '设施高程': 'elevation',
                '直接联系人': 'link_man',
                '安全联系人': 'safety_man',
                '安全联系人电话': 'safety_tel',
                '经度': 'lng',
                '纬度': 'lat',
                '联系人电话': 'link_tel',
                '详细地址': 'address'
            }
        },
        '学校': {
            'filter_type': [1],
            'columns': {
                '区县级名称': 'addvnm',
                '区县级区划编码': 'addvcd',
                '乡镇编码': 'region_code',
                '乡镇名称': 'region_name',
                '设施编码': 'ssid',
                '重要设施名称': 'ssnm',
                '保护片编码': 'polder_code',
                '详细地址': 'address',
                '联系人': 'link_man',
                '联系方式': 'link_tel',
                '设施高程': 'elevation',
                '经度': 'lng',
                '纬度': 'lat'
            }
        },
        '公园景点': {
            'filter_type': [7],
            'columns': {
                '区县级名称': 'addvnm',
                '区县级区划编码': 'addvcd',
                '乡镇编码': 'region_code',
                '乡镇名称': 'region_name',
                '设施id': 'ssid',
                '重要设施名称': 'ssnm',
                '保护片编码': 'polder_code',
                '详细地址': 'address',
                '直接联系人': 'link_man',
                '联系人电话': 'link_tel',
                '设施高程': 'elevation',
                '经度': 'lng',
                '纬度': 'lat'
            }
        },
        '敬老院': {
            'filter_type': [5],
            'columns': {
                '区县级名称': 'addvnm',
                '区县级区划编码': 'addvcd',
                '乡镇编码': 'region_code',
                '乡镇名称': 'region_name',
                '设施编码': 'ssid',
                '重要设施名称': 'ssnm',
                '保护片编码': 'polder_code',
                '详细地址': 'address',
                '联系人': 'link_man',
                '联系方式': 'link_tel',
                '设施高程': 'elevation',
                '经度': 'lng',
                '纬度': 'lat'
            }
        },
        '电站': {
            'filter_type': [8],
            'columns': {
                '区县级名称': 'addvnm',
                '区县级区划编码': 'addvcd',
                '乡镇编码': 'region_code',
                '乡镇名称': 'region_name',
                '重要设施编码': 'ssid',
                '重要设施名称': 'ssnm',
                '保护片编码': 'polder_code',
                '详细地址': 'address',
                '设施高程': 'elevation',
                '联系人': 'link_man',
                '联系电话': 'link_tel',
                '管理单位': 'management_unit'
            }
        },
        '水厂': {
            'filter_type': [3],
            'columns': {
                '区县级名称': 'addvnm',
                '区县级区划编码': 'addvcd',
                '乡镇编码': 'region_code',
                '设施id': 'ssid',
                '乡镇名称': 'region_name',
                '重要设施名称': 'ssnm',
                '详细地址': 'address',
                '保护片编码': 'polder_code',
                '联系人电话': 'link_tel',
                '直接联系人': 'link_man',
                '设施高程': 'elevation',
                '经度': 'lng',
                '纬度': 'lat'
            }
        },
        '政府机构': {
            'filter_type': [10],
            'columns': {
                '区县级名称': 'addvnm',
                '区县级区划编码': 'addvcd',
                '乡镇编码': 'region_code',
                '乡镇名称': 'region_name',
                '设施编码': 'ssid',
                '重要设施名称': 'ssnm',
                '保护片编码': 'polder_code',
                '详细地址': 'address',
                '联系人': 'link_man',
                '联系方式': 'link_tel',
                '设施高程': 'elevation',
                '经度': 'lng',
                '纬度': 'lat'
            }
        },
        '重要设施': {
            'filter_type': [0, 1, 2, 5, 7, 10],  # 所有有数据的类型
            'columns': {
                '重要设施名称': 'fac_name',
                '重要设施编码': 'fac_code',
                '设施类型': 'fac_type',
                '设施所处网格': 'grid_code',
                '高程': 'elevation',
                '所处保护片编码': 'polder_code'
            }
        }
    }
    
    print("\n" + "="*60)
    print("开始填充设施数据")
    print("="*60)
    
    # 加载 Excel
    try:
        wb = load_workbook(excel_path)
        print(f"✅ 成功打开 Excel: {excel_path}")
    except Exception as e:
        print(f"❌ 打开 Excel 失败: {e}")
        return
    
    # 统计信息
    stats = {}
    
    # 遍历每个 sheet 进行填充
    for sheet_name, config in SHEET_CONFIGS.items():
        print(f"\n{'='*60}")
        print(f"📝 处理 sheet: {sheet_name}")
        print(f"{'='*60}")
        
        # 检查 sheet 是否存在
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  警告: sheet '{sheet_name}' 不存在，跳过")
            continue
        
        ws = wb[sheet_name]
        
        # 清空原有数据（保留表头，表头在第2行）
        if ws.max_row > 2:
            ws.delete_rows(3, ws.max_row - 2)
            print(f"🧹 已清空原有数据")
        
        # 构建列名到列号的映射（表头在第2行）
        header_row = 2
        column_mapping = {}
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value:
                column_mapping[cell_value] = col
        
        print(f"📋 表头映射: {list(column_mapping.keys())}")
        
        # 筛选对应类型的设施
        filter_types = config['filter_type']
        columns = config['columns']
        
        filled_count = 0
        current_row = 3  # 数据从第3行开始（第1行说明，第2行表头）
        
        for feature in geojson_data['features']:
            props = feature.get('properties', {})
            
            # 筛选类型
            fac_type = props.get('type')
            if fac_type not in filter_types:
                continue
            
            # 提取数据
            county_name = props.get('县（区', '')  # 武义县
            city_name = props.get('地级市', '')  # 金华市
            town_name = props.get('town', '')  # 乡镇名称
            fac_name = props.get('名称', '')
            fac_code = props.get('fac_code', '')
            address = props.get('地址', '')
            elevation = props.get('高程')
            lng = props.get('经度')
            lat = props.get('纬度')
            link_tel = props.get('联系电', '')
            polder_id = normalize_code(props.get('polderId', ''))
            grid_id = props.get('grid_id')
            
            # 获取区划编码
            county_code = get_county_code(county_name, region_codes)
            town_code = get_town_code(town_name, region_codes)
            
            # 填充数据
            for col_name, field_key in columns.items():
                if col_name not in column_mapping:
                    continue
                
                col_idx = column_mapping[col_name]
                
                # 根据字段键设置值
                if field_key == 'addvnm':
                    value = county_name
                elif field_key == 'addvcd':
                    value = county_code
                elif field_key == 'region_code':
                    value = town_code
                elif field_key == 'region_name':
                    value = town_name
                elif field_key == 'ssid':
                    value = fac_code
                elif field_key == 'ssnm':
                    value = fac_name
                elif field_key == 'polder_code':
                    value = polder_id.upper()
                elif field_key == 'address':
                    value = address
                elif field_key == 'elevation':
                    value = elevation
                elif field_key == 'lng':
                    value = lng
                elif field_key == 'lat':
                    value = lat
                elif field_key == 'link_man':
                    value = ''  # GeoJSON 中没有联系人字段
                elif field_key == 'link_tel':
                    value = link_tel
                elif field_key == 'fac_name':
                    value = fac_name
                elif field_key == 'fac_code':
                    value = fac_code
                elif field_key == 'fac_type':
                    value = fac_type
                elif field_key == 'grid_code':
                    value = grid_id
                # 其他空字段
                elif field_key in ['legal_person', 'legal_tel', 'safety_man', 'safety_tel', 
                                    'station_type', 'management_unit']:
                    value = ''  # 这些字段在 GeoJSON 中没有
                else:
                    value = ''
                
                ws.cell(row=current_row, column=col_idx).value = value
            
            current_row += 1
            filled_count += 1
            
            # 进度显示
            if filled_count % 10 == 0:
                print(f"  ⏳ 已填充 {filled_count} 条数据...")
        
        stats[sheet_name] = filled_count
        print(f"✅ {sheet_name}: 填充完成，共 {filled_count} 条数据")
    
    # 保存 Excel
    try:
        wb.save(excel_path)
        print(f"\n💾 已保存 Excel: {excel_path}")
    except Exception as e:
        print(f"❌ 保存 Excel 失败: {e}")
        return
    
    # 生成报告
    generate_report(stats, excel_path)

def generate_report(stats, excel_path):
    """生成填充报告"""
    report_path = 'output/设施填充报告.txt'
    
    try:
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write("="*60 + "\n")
            f.write("设施数据填充报告\n")
            f.write("="*60 + "\n\n")
            f.write(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"目标文件: {excel_path}\n\n")
            
            f.write("填充统计:\n")
            f.write("-"*60 + "\n")
            total_count = 0
            for sheet_name, count in stats.items():
                f.write(f"  {sheet_name:20s}: {count:5d} 条\n")
                total_count += count
            f.write("-"*60 + "\n")
            f.write(f"  {'总计':20s}: {total_count:5d} 条\n\n")
            
            f.write("填充完成！\n")
        
        print(f"\n📊 报告已生成: {report_path}")
    
    except Exception as e:
        print(f"❌ 生成报告失败: {e}")

# ============ 主函数 ============

def main():
    """主函数"""
    parser = argparse.ArgumentParser(
        description='设施信息填充',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例: python3 3.09_risk_facilities.py -g /path/to/baohu.geojson -e /path/to/hx.xlsx'
    )
    parser.add_argument('-g', '--geojson', help='GeoJSON 文件路径 (baohu.geojson)')
    parser.add_argument('-e', '--excel', help='Excel 文件路径 (risk_xx.xlsx)')
    parser.add_argument('-r', '--region', help='乡镇编码对照表路径')
    args = parser.parse_args()
    
    geojson_path = args.geojson or 'input/baohu/baohu.geojson'
    region_codes_path = args.region or 'input/region_name_code.csv'
    excel_path = args.excel or 'output/risk_sx.xlsx'
    
    print("\n" + "="*60)
    print("设施数据填充工具 (v2.0)")
    print("="*60)
    print(f"  GeoJSON: {geojson_path}")
    print(f"  Excel:   {excel_path}")
    print("="*60 + "\n")
    
    # 1. 加载数据
    geojson_data = load_geojson(geojson_path)
    if geojson_data is None:
        return
    
    region_codes = load_region_codes(region_codes_path)
    
    # 2. 检查 Excel 文件
    if not os.path.exists(excel_path):
        print(f"❌ 错误: Excel 文件不存在 - {excel_path}")
        return
    
    # 3. 创建备份
    create_backup(excel_path)
    
    # 4. 填充数据
    fill_facilities_data(geojson_data, region_codes, excel_path)
    
    print("\n" + "="*60)
    print("✅ 所有设施数据填充完成！")
    print("="*60 + "\n")

if __name__ == "__main__":
    main()

