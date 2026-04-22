#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

excel_path = 'output/risk_sx.xlsx'
wb = load_workbook(excel_path)

# 检查几个关键 sheet 的表头
sheets_to_check = ['学校', '医院', '重要设施']

for sheet_name in sheets_to_check:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n{"="*80}')
        print(f'Sheet: {sheet_name}')
        print(f'{"="*80}')
        print(f'总行数: {ws.max_row}, 总列数: {ws.max_column}')
        print('\n表头（第1行）所有列名:')
        print('-'*80)
        
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=1, column=col_idx).value
            print(f'  列{col_idx:2d}: {repr(cell_value)}')
        
        # 检查第2行（如果有）
        if ws.max_row >= 2:
            print('\n第2行内容:')
            print('-'*80)
            for col_idx in range(1, min(6, ws.max_column + 1)):
                cell_value = ws.cell(row=2, column=col_idx).value
                print(f'  列{col_idx}: {cell_value}')
        else:
            print('\n⚠️  没有数据行！')

