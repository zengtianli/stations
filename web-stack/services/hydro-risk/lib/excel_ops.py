#!/usr/bin/env python3
"""
Excel 操作工具 - 统一的 Excel/CSV 数据读写功能

提供 sheet 级别的数据读写操作，与 core/xlsx_core.py（格式转换）互补。
"""

import sys

import pandas as pd
from openpyxl import load_workbook


def load_excel_sheet(excel_path, sheet_name):
    """
    加载 Excel 文件的指定 sheet

    Args:
        excel_path: Excel 文件路径
        sheet_name: sheet 名称

    Returns:
        DataFrame: pandas DataFrame 对象

    Raises:
        SystemExit: 文件或 sheet 不存在时退出
    """
    try:
        wb = load_workbook(excel_path)
        if sheet_name not in wb.sheetnames:
            print(f"✗ 错误: Excel文件中没有名为 '{sheet_name}' 的sheet")
            print(f"  可用的sheets: {wb.sheetnames}")
            sys.exit(1)

        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        print(f"✓ 成功加载 {excel_path} 的 '{sheet_name}' sheet")
        print(f"  包含 {len(df)} 行数据")
        return df

    except FileNotFoundError:
        print(f"✗ 错误: 找不到Excel文件 {excel_path}")
        sys.exit(1)
    except Exception as e:
        print(f"✗ 错误: 加载Excel失败 - {e}")
        sys.exit(1)


def save_to_excel(df, excel_path, sheet_name, index=False):
    """
    将 DataFrame 保存到 Excel 文件

    Args:
        df: pandas DataFrame
        excel_path: Excel 文件路径
        sheet_name: sheet 名称
        index: 是否保存索引
    """
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=index)
        print(f"✓ 已保存数据到 {excel_path} 的 '{sheet_name}' sheet")
        print(f"  共 {len(df)} 行数据")
    except Exception as e:
        print(f"✗ 错误: 保存Excel失败 - {e}")
        sys.exit(1)


def load_csv_data(csv_path, dtype=None):
    """
    读取 CSV 数据

    Args:
        csv_path: CSV 文件路径
        dtype: 数据类型字典

    Returns:
        DataFrame: pandas DataFrame 对象
    """
    try:
        if dtype:
            df = pd.read_csv(csv_path, encoding='utf-8', dtype=dtype)
        else:
            df = pd.read_csv(csv_path, encoding='utf-8')
        print(f"✓ 成功读取 CSV 文件: {csv_path}")
        print(f"  包含 {len(df)} 条记录")
        return df
    except Exception as e:
        print(f"⚠ 警告: 无法读取CSV文件 {csv_path} - {e}")
        return pd.DataFrame()


def check_columns_exist(df, required_columns, data_desc="数据"):
    """
    检查 DataFrame 是否包含必需的列

    Args:
        df: pandas DataFrame
        required_columns: 必需的列名列表
        data_desc: 数据描述（用于错误信息）

    Raises:
        SystemExit: 缺少必需列时退出
    """
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        print(f"✗ 错误: {data_desc}缺少必需的列:")
        for col in missing_columns:
            print(f"  - {col}")
        print(f"\n  当前列: {list(df.columns)}")
        sys.exit(1)
