#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

excel_path = 'output/risk_sx.xlsx'
wb = load_workbook(excel_path)

# 检查关键 sheet 的数据行数
sheets_to_check = [
    '保护片信息 ',
    '保护片行政区域信息  ',
    '保护片堤段对应关系信息',
    '堤段对应信息',
    '保护片堤段堤顶高程对应关系信息',
    '断面里程对应关系信息',
    '堤防信息',
    '堤防纵剖面数据',
    '学校',
    '医院',
    '重要设施'
]

print('='*80)
print('risk_sx.xlsx 数据行数检查')
print('='*80)

for sheet_name in sheets_to_check:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data_rows = ws.max_row - 2  # 减去说明行和表头行
        print(f'{sheet_name:40s}: {data_rows:4d} 行')
    else:
        print(f'{sheet_name:40s}: ⚠️  不存在')

print('='*80)

